"""
analytics views - 关键词权重 & 日报 & 统计 & 系统指标 & 组织报表 & 队列监控
"""
import io
import csv

from loguru import logger
from datetime import timedelta, datetime

from django.db import models
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.analytics.models import (
    KeywordWeight, SystemMetricsReport,
    OrgUsageReport, ChunkClickLog, KeywordFeedbackAgg,
)
from apps.chat.models import QaRecord, QaFeedback
from apps.users.permissions import CanViewAnalytics
from rag_project.config import AnalyticsConfig


# ============================================================================
# 组织筛选工具函数
# ============================================================================

def _parse_org_scope(request):
    """从 request.query_params 解析 dept_id/team_id,返回 (dept_id, team_id)。

    team_id 有值时 dept_id 自动忽略(团队天然属于某部门,过滤更精确);
    dept_id 有值时用 user__department_id=dept_id 过滤(包含部门所有团队成员);
    两者都为空返回 (None, None),调用方跳过组织过滤。
    """
    dept_id = request.query_params.get('dept_id', '').strip() or None
    team_id = request.query_params.get('team_id', '').strip() or None
    if dept_id:
        try:
            dept_id = int(dept_id)
        except (ValueError, TypeError):
            dept_id = None
    if team_id:
        try:
            team_id = int(team_id)
        except (ValueError, TypeError):
            team_id = None
    return dept_id, team_id


def _apply_org_filter_on_qa(qs, dept_id, team_id, qa_prefix=''):
    """对以 QaRecord 为 JOIN 起点的 QuerySet 应用组织筛选(按提问用户归属)。

    qa_prefix: 当 QS 是 JOIN 后的表时(如 MultiDimensionScore),传入 qa 关联前缀
    (如 'qa_record__'),最终生成 qa_record__user__department_id。空串表示 qs 就是 QaRecord。
    前缀必须以 '__' 结尾(或为空串),否则会拼出错误的 ORM lookup 路径。
    """
    # 统一规范化前缀:非空时确保以 '__' 结尾,再拼接 'user__'
    base = (qa_prefix + '__') if (qa_prefix and not qa_prefix.endswith('__')) else qa_prefix
    base += 'user__'
    # team 有值时直接按团队过滤,更精确,无需再按部门过滤
    if team_id:
        return qs.filter(**{f'{base}team_id': team_id})
    if dept_id:
        return qs.filter(**{f'{base}department_id': dept_id})
    return qs


def _apply_org_filter_on_doc(qs, dept_id, team_id, doc_prefix=''):
    """对以 Document 为 JOIN 起点的 QuerySet 应用组织筛选(按文档 dept_id/team_id 归属)。

    doc_prefix: JOIN 前缀(如 'document__'),空串表示 qs 本身就是 Document。
    """
    # team 有值时直接按团队过滤(团队归属文档或冗余 dept_id 已对齐的团队文档)
    if team_id:
        return qs.filter(**{f'{doc_prefix}team_id': team_id})
    if dept_id:
        # 部门级:直接归属部门(dept_id=X,team_id 空)或其下属团队的文档(dept_id 冗余对齐)
        return qs.filter(**{f'{doc_prefix}dept_id': dept_id})
    return qs


class KeywordWeightListView(APIView):
    """GET /api/v1/analytics/keywords/?root_type=&top=20

    - 关键词权重列表，管理员可查看
    - top 参数限制 1-500，防止恶意拉取全表
    - POST 时校验 keyword 非空 + weight_score 范围 0.1~5.0
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request):
        try:
            top = int(request.query_params.get("top") or 20)
        except (ValueError, TypeError):
            return Response({"detail": "top 必须为整数"}, status=400)
        # top 上限 500，防止 OOM / 全表扫描
        top = max(1, min(top, 500))
        root_type = request.query_params.get("root_type")
        qs = KeywordWeight.objects.all().order_by("-weight_score")
        if root_type:
            qs = qs.filter(root_type=root_type)
        rows = list(qs[:top].values(
            "id", "keyword", "root_type", "hit_count", "good_feedback", "bad_feedback",
            "weight_score", "updated_at"
        ))
        return Response({"rows": rows, "count": len(rows)})

    def post(self, request):
        keyword = (request.data.get("keyword") or "").strip()
        weight_score = request.data.get("weight_score")
        # root_type 长度与模型 CharField(max_length=32) 对齐；
        # 超长时在 API 层直接返回 400，避免 DB 层 DataError 或静默截断
        root_type = (request.data.get("root_type", "all") or "").strip() or "all"
        if not keyword:
            return Response({"detail": "keyword 必填"}, status=400)
        if len(keyword) > 64:
            return Response({"detail": "keyword 长度不能超过 64 字符"}, status=400)
        if len(root_type) > 32:
            return Response({"detail": "root_type 长度不能超过 32 字符"}, status=400)
        # weight_score 防御性钳位：未传默认 1.0，范围 0.1~5.0
        try:
            weight_score = float(weight_score) if weight_score is not None else 1.0
        except (TypeError, ValueError):
            return Response({"detail": "weight_score 必须为数字"}, status=400)
        weight_score = max(0.1, min(5.0, weight_score))

        obj, created = KeywordWeight.objects.update_or_create(
            keyword=keyword, root_type=root_type,
            defaults={"weight_score": weight_score}
        )
        return Response({
            "id": obj.id, "keyword": obj.keyword, "root_type": obj.root_type,
            "weight_score": obj.weight_score, "created": created
        })


class KeywordWeightDetailView(APIView):
    """PUT /api/v1/analytics/keywords/{id}/ 调整权重

    - 修改关键词权重直接影响检索排序（weight_score 范围 0.1~2.0）
    - 仅限管理员操作，需记录审计日志
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.write'

    def put(self, request, kw_id):
        delta = request.data.get("delta")
        if delta is None:
            return Response({"detail": "delta 必填"}, status=400)
        try:
            delta = float(delta)
        except (TypeError, ValueError):
            return Response({"detail": "delta 必须为数字"}, status=400)

        try:
            kw = KeywordWeight.objects.get(id=kw_id)
            old_score = kw.weight_score
            kw.weight_score = max(0.1, min(2.0, kw.weight_score + delta))
            kw.save(update_fields=["weight_score"])
            # 手动调整同样落 KeywordFeedbackAgg，与自动调整统一审计展示；
            # 同 (日期, 关键词) 自动任务将跳过应用，保证手动覆盖优先
            from apps.analytics.feedback_loop import record_manual_adjustment
            record_manual_adjustment(kw, old_score, request.user)
            logger.info(
                f"keyword_weight_adjusted kw_id={kw.id} keyword={kw.keyword} old={old_score:.2f} delta={delta:+.2f} new={kw.weight_score:.2f} user={request.user.username}"
            )
            return Response({
                "id": kw.id, "keyword": kw.keyword, "weight_score": kw.weight_score
            })
        except KeywordWeight.DoesNotExist:
            return Response({"detail": "keyword weight 不存在"}, status=404)


class ChunkClickLogView(APIView):
    """POST /api/v1/analytics/chunk-clicks/ - 记录一次溯源来源点击

    前端在用户点击回答的引用卡片时调用（聊天页 previewCitation 埋点），
    作为反馈闭环"点击率"的原始数据源。点击是低频行为，直接 INSERT 落库。

    请求体：
        chunk_id: int 必填（被点击的 chunk id）
        qa_record_id: int 可选（本次回答的 QaRecord.id，聚合点击率时需要）
        document_id: int 可选
        root_type: str 可选（默认 all）
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            chunk_id = int(request.data.get("chunk_id"))
        except (TypeError, ValueError):
            return Response({"detail": "chunk_id 必填且为整数"}, status=400)
        if chunk_id <= 0:
            return Response({"detail": "chunk_id 必须为正整数"}, status=400)

        qa_record_id = request.data.get("qa_record_id")
        try:
            qa_record_id = int(qa_record_id) if qa_record_id not in (None, "") else None
        except (TypeError, ValueError):
            qa_record_id = None
        document_id = request.data.get("document_id")
        try:
            document_id = int(document_id) if document_id not in (None, "") else None
        except (TypeError, ValueError):
            document_id = None
        root_type = (request.data.get("root_type") or "all").strip() or "all"

        # 记录点击日志；qa_record/user 以请求携带为准，写入失败不阻塞前端
        ChunkClickLog.objects.create(
            user=request.user,
            qa_record_id=qa_record_id,
            document_id=document_id,
            chunk_id=chunk_id,
            root_type=root_type,
        )
        return Response({"ok": True})


class KeywordFeedbackAggListView(APIView):
    """GET /api/v1/analytics/feedback-loop/aggregations/ - 自动调整记录列表

    支持 date/keyword/root_type/status 筛选，limit 上限 100，
    供运营工具展示反馈闭环的每日聚合结果与权重调整审计。
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request):
        qs = KeywordFeedbackAgg.objects.all()
        date = (request.query_params.get("date") or "").strip()
        if date:
            try:
                qs = qs.filter(report_date=date)
            except ValueError:
                return Response({"detail": "date 格式应为 YYYY-MM-DD"}, status=400)
        keyword = (request.query_params.get("keyword") or "").strip()
        if keyword:
            qs = qs.filter(keyword__icontains=keyword)
        root_type = (request.query_params.get("root_type") or "").strip()
        if root_type:
            qs = qs.filter(root_type=root_type)
        status = (request.query_params.get("status") or "").strip()
        if status:
            qs = qs.filter(status=status)
        try:
            limit = max(1, min(int(request.query_params.get("limit") or 50), 100))
        except (ValueError, TypeError):
            limit = 50

        rows = list(qs.order_by("-report_date", "-created_at")[:limit].values(
            "id", "report_date", "keyword", "root_type",
            "shown_count", "click_count", "adopt_count", "bad_count",
            "click_rate", "adopt_rate",
            "old_score", "new_score", "delta", "reason",
            "adjust_type", "status", "applied_at", "created_at",
        ))
        return Response({"rows": rows, "count": len(rows)})


class KeywordFeedbackApplyView(APIView):
    """POST /api/v1/analytics/feedback-loop/apply/ - 人工复核应用/忽略待调整记录

    请求体：{id: int, action: 'apply'|'ignore'}
    人工复核开关(AUTO_APPLY=False)产生的 pending 记录由运营在此确认后生效。
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.write'

    def post(self, request):
        from apps.analytics.feedback_loop import apply_pending_adjustment
        try:
            agg_id = int(request.data.get("id"))
        except (TypeError, ValueError):
            return Response({"detail": "id 必填且为整数"}, status=400)
        action = (request.data.get("action") or "apply").strip()
        if action not in ("apply", "ignore"):
            return Response({"detail": "action 必须为 apply 或 ignore"}, status=400)

        ok, message = apply_pending_adjustment(agg_id, action=action, user=request.user)
        if not ok:
            return Response({"detail": message}, status=400)
        return Response({"ok": True, "detail": message})


class RunFeedbackLoopView(APIView):
    """POST /api/v1/analytics/feedback-loop/run/ - 手动触发反馈闭环聚合

    请求体：{date: 'YYYY-MM-DD' 可选}，默认聚合昨天。
    与每日定时任务共用同一逻辑，支持回补指定日期。
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.write'

    def post(self, request):
        from apps.analytics.feedback_loop import run_keyword_feedback_loop
        date = request.data.get("date") or None
        try:
            result = run_keyword_feedback_loop(report_date=date)
        except ValueError as e:
            return Response({"detail": str(e)}, status=400)
        return Response(result)


class DailyReportView(APIView):
    """GET /api/v1/analytics/daily/ 日报
    - 展示最近 2 天（今日/昨日）的 QA 概览 + 反馈统计
    - 使用条件聚合一次性查出 qa_count/good/bad，避免 3 次 COUNT 查询
    - 仅聚合实时数据（T+1 精确指标请使用 SystemMetricsReport）
    - 支持 root_type 参数（领域过滤），与前端 loadDailyReport 传参对齐
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request):
        from django.db.models.functions import TruncDate

        # 本地业务日期而非 UTC 日期：timezone.now().date() 在凌晨时段与 PG
        # __date/TruncDate 的本地时区转换相差一天，导致今日数据统计错天（与 Trend 视图一致）
        today = timezone.localdate()
        yesterday = today - timedelta(days=1)
        root_type = request.query_params.get("root_type")

        # —— 1. 按日条件聚合 QaRecord 数量（一次查询出 2 天的 qa_count）——
        qa_qs = (QaRecord.objects
                 .filter(created_at__date__in=[today, yesterday]))
        if root_type:
            qa_qs = qa_qs.filter(root_type=root_type)
        qa_qs = (qa_qs
                 .annotate(day=TruncDate('created_at'))
                 .values('day')
                 .annotate(qa_count=models.Count('id')))
        qa_map = {str(r['day']): r['qa_count'] for r in qa_qs}

        # —— 2. 按日条件聚合 好评/差评（一次查询出 2 天，good+bad 同时出）——
        fb_qs = (QaFeedback.objects
                 .filter(qa_record__created_at__date__in=[today, yesterday]))
        if root_type:
            fb_qs = fb_qs.filter(qa_record__root_type=root_type)
        fb_qs = (fb_qs
                 .annotate(day=TruncDate('qa_record__created_at'))
                 .values('day')
                 .annotate(
                     good=models.Count('id', filter=models.Q(rating__gt=0)),
                     bad=models.Count('id', filter=models.Q(rating__lt=0)),
                 ))
        fb_map = {str(r['day']): {'good': r['good'] or 0, 'bad': r['bad'] or 0} for r in fb_qs}

        result = {"today": None, "yesterday": None}
        for key, day in [("today", today), ("yesterday", yesterday)]:
            day_str = day.isoformat()
            qa_count = qa_map.get(day_str, 0)
            fb = fb_map.get(day_str, {'good': 0, 'bad': 0})
            good, bad = fb['good'], fb['bad']
            result[key] = {
                "date": day_str,
                "qa_count": qa_count,
                "good": good,
                "bad": bad,
                "accuracy": round(good / max(good + bad, 1), 4),
            }

        return Response(result)


class TrendReportView(APIView):
    """GET /api/v1/analytics/trend/?days=7&root_type=
    或 GET /api/v1/analytics/trend/?start_date=2025-01-01&end_date=2025-01-15&root_type=
    - 按日汇总 QA 次数、Token、费用、准确率等趋势
    - 系统级统计，需具备 analytics:system:read 权限
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request):
        from django.db.models.functions import TruncDate

        start_date_param = request.query_params.get("start_date")
        end_date_param = request.query_params.get("end_date")
        root_type = request.query_params.get("root_type")

        if start_date_param and end_date_param:
            try:
                start_date = datetime.fromisoformat(start_date_param).date()
                end_date = datetime.fromisoformat(end_date_param).date()
            except ValueError:
                return Response({"detail": "日期格式应为 YYYY-MM-DD"}, status=400)
            if start_date > end_date:
                return Response({"detail": "开始日期不能晚于结束日期"}, status=400)
            if (end_date - start_date).days > 365:
                return Response({"detail": "自定义范围最多 365 天"}, status=400)
        else:
            # days 参数防御：非整数返回 400，范围限制 1-365 防止过大循环
            try:
                days = int(request.query_params.get("days") or 7)
            except (ValueError, TypeError):
                return Response({"detail": "days 必须为整数"}, status=400)
            if days < 1 or days > 365:
                return Response({"detail": "days 范围应为 1-365"}, status=400)
            end_date = timezone.localdate()
            start_date = end_date - timedelta(days=days - 1)

        # 构造 QaRecord 基础过滤 QS，先不要调用 values()，方便后续复用两次聚合
        qa_base_qs = QaRecord.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date,
        )
        if root_type:
            qa_base_qs = qa_base_qs.filter(root_type=root_type)

        # 聚合 1：qa count（全量，包含缓存命中）
        qa_counts = (qa_base_qs
                      .annotate(day=TruncDate("created_at"))
                      .values("day")
                      .annotate(qa_count=models.Count("id"))
                      .order_by("day"))
        qa_map = {r["day"].isoformat(): {"qa_count": r["qa_count"]} for r in qa_counts}

        # 聚合 2：仅非缓存命中的平均耗时，分别计算首字(TTFT)和整体(total)耗时
        # —— 缓存命中请求走旁路，没有真实 LLM 生成过程，latency 代表纯读库速度，会拉低真实生成耗时
        latency_by_day = (qa_base_qs
                           .exclude(is_hit_cache=True)
                           .annotate(day=TruncDate("created_at"))
                           .values("day")
                           .annotate(avg_total=models.Avg("latency_total_ms"),
                                      avg_ttfb=models.Avg("latency_ttfb_ms"))
                           .order_by("day"))
        lat_map = {r["day"].isoformat(): r for r in latency_by_day}

        # 单次 GROUP BY 查询：good / bad feedback count
        # —— 注意：所有 filter（尤其是 JOIN 条件如 qa_record__root_type）必须放在 .values().annotate() 之前 ——
        #    否则如果 GROUP BY 已经执行完，再加 JOIN filter 会导致 Django 生成 subquery，
        #    不同版本行为不一致。提前 filter 也能减少 GROUP BY 需要处理的行数。
        fb_base = QaFeedback.objects.filter(
            qa_record__created_at__date__gte=start_date,
            qa_record__created_at__date__lte=end_date,
        )
        if root_type:
            fb_base = fb_base.filter(qa_record__root_type=root_type)
        fb_qs = (fb_base
                 .annotate(day=TruncDate("qa_record__created_at"))
                 .values("day")
                 .annotate(good=models.Count("id", filter=models.Q(rating__gt=0)),
                           bad=models.Count("id", filter=models.Q(rating__lt=0)))
                 .order_by("day"))
        fb_map = {r["day"].isoformat(): r for r in fb_qs}

        days_total = (end_date - start_date).days + 1
        trend = []
        for i in range(days_total):
            day = start_date + timedelta(days=i)
            day_str = day.isoformat()
            qa_row = qa_map.get(day_str, {})
            lat_row = lat_map.get(day_str, {})
            fb_row = fb_map.get(day_str, {})
            qa_count = qa_row.get("qa_count", 0)
            good = fb_row.get("good", 0)
            bad = fb_row.get("bad", 0)
            trend.append({
                "date": day_str,
                "qa_count": qa_count,
                "good": good,
                "bad": bad,
                "accuracy": round(good / max(good + bad, 1), 4),
                # 仅非缓存命中的真实生成耗时，缓存命中延迟仅代表纯读库速度不计入
                "avg_ttft_ms": int(lat_row.get("avg_ttfb", 0) or 0),
                "avg_total_ms": int(lat_row.get("avg_total", 0) or 0),
            })

        return Response({"trend": trend, "days": len(trend)})


class BadFeedbackListView(APIView):
    """GET /api/v1/analytics/bad-feedbacks/?top=20&root_type=

    - 差评反馈列表，仅管理员可查看
    - top 参数限制 1-500，防止恶意拉取全表
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request):
        try:
            top = int(request.query_params.get("top") or 20)
        except (ValueError, TypeError):
            return Response({"detail": "top 必须为整数"}, status=400)
        # top 上限 500，防止 OOM
        top = max(1, min(top, 500))
        root_type = request.query_params.get("root_type")
        qs = QaFeedback.objects.filter(rating__lt=0).select_related("qa_record", "user")
        if root_type:
            qs = qs.filter(qa_record__root_type=root_type)
        qs = qs.order_by("-created_at")[:top]
        rows = []
        for fb in qs:
            # question/answer 防御性安全取值：字段本身非空时才切片，None 或空字符串安全降级
            question_val = fb.qa_record.question if (fb.qa_record and fb.qa_record.question is not None) else ""
            answer_val = fb.qa_record.answer if (fb.qa_record and fb.qa_record.answer is not None) else ""
            rows.append({
                "id": fb.id,
                "qa_record_id": fb.qa_record_id,
                "question": question_val[:100],
                "answer": answer_val[:100],
                "rating": fb.rating,
                "tags": fb.tags,
                "comment": fb.comment,
                "status": fb.status,
                "user": fb.user.real_name or fb.user.username if fb.user else "",
                "created_at": fb.created_at.isoformat() if fb.created_at else "",
            })
        return Response({"rows": rows, "count": len(rows)})


class OverviewStatsView(APIView):
    """GET /api/v1/analytics/overview/?root_type= 概览统计
    - Dashboard 顶部卡片：总 QA 次数、近 7 日 QA、准确率、平均延迟、活跃用户、文档数、活跃会话
    - 使用条件聚合一次性查出 total_qa/weekly_qa/good/bad，减少查询
    - 系统级统计，需具备 analytics:system:read 权限
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request):
        from apps.knowledge.models import Document
        from apps.memory.models import Session

        today = timezone.now().date()
        week_ago = today - timedelta(days=7)
        root_type = request.query_params.get("root_type")

        qa_qs = QaRecord.objects.all()
        if root_type:
            qa_qs = qa_qs.filter(root_type=root_type)

        # 一次性聚合：总 QA + 近 7 日 QA
        qa_agg = qa_qs.aggregate(
            total_qa=models.Count('id'),
            weekly_qa=models.Count('id', filter=models.Q(created_at__date__gte=week_ago)),
        )
        total_qa = qa_agg['total_qa'] or 0
        weekly_qa = qa_agg['weekly_qa'] or 0

        # 反馈聚合：一次性查出好评+差评
        fb_qs = QaFeedback.objects.all()
        if root_type:
            fb_qs = fb_qs.filter(qa_record__root_type=root_type)
        fb_agg = fb_qs.aggregate(
            good_fb=models.Count('id', filter=models.Q(rating__gt=0)),
            bad_fb=models.Count('id', filter=models.Q(rating__lt=0)),
        )
        good_fb = fb_agg['good_fb'] or 0
        bad_fb = fb_agg['bad_fb'] or 0
        accuracy = round(good_fb / max(good_fb + bad_fb, 1), 4)

        # 平均延迟：排除缓存命中（缓存命中延迟仅代表纯读库速度，不是真实 LLM 生成耗时）
        norm_qs = qa_qs.exclude(is_hit_cache=True)
        latency_agg = norm_qs.aggregate(
            avg_total=models.Avg("latency_total_ms"),
            avg_ttfb=models.Avg("latency_ttfb_ms"),
        )
        avg_total_ms = int(latency_agg["avg_total"] or 0)
        avg_ttft_ms = int(latency_agg["avg_ttfb"] or 0)

        active_users = qa_qs.filter(created_at__date__gte=week_ago).values("user").distinct().count()

        # 文档统计：一次性查出 total + completed
        doc_qs = Document.objects.filter(is_deleted=False)
        if root_type:
            doc_qs = doc_qs.filter(root_type=root_type)
        doc_agg = doc_qs.aggregate(
            total_docs=models.Count('id'),
            completed_docs=models.Count('id', filter=models.Q(status='done')),
        )
        total_docs = doc_agg['total_docs'] or 0
        completed_docs = doc_agg['completed_docs'] or 0

        active_sessions = Session.objects.filter(is_deleted=False, last_active_at__date__gte=week_ago).count()

        return Response({
            "total_qa": total_qa,
            "weekly_qa": weekly_qa,
            "accuracy": accuracy,
            "avg_latency_ms": avg_total_ms,   # 兼容旧字段，等于 avg_total_ms
            "avg_ttft_ms": avg_ttft_ms,      # 首字耗时（排除缓存命中）
            "avg_total_ms": avg_total_ms,     # 整体总耗时（排除缓存命中）
            "active_users": active_users,
            "total_docs": total_docs,
            "completed_docs": completed_docs,
            "active_sessions": active_sessions,
        })


class QaRecordView(APIView):
    """GET /api/v1/analytics/qa-records/?start_date=&end_date=&root_type=&qa_id=

    - 返回 QA 记录列表，供 Dashboard 查看对话历史
    - 可选 qa_id 参数：返回单条 QA 详情（用于 QA 详情弹窗，避免列表前 100 条限制）
    - 仅管理员或具备 analytics:system:read 权限的用户可访问
    - 支持日期范围、领域过滤和分页
    - feedback 是 OneToOne，需用 hasattr 判断存在性（select_related 做 LEFT JOIN，无反馈时为 None）
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request):

        # —— 单条详情：qa_id 参数，用于详情弹窗（绕开分页前 100 条的限制）——
        qa_id = request.query_params.get('qa_id')
        if qa_id:
            try:
                qa_id = int(qa_id)
            except (ValueError, TypeError):
                return Response({'detail': 'qa_id 必须为整数'}, status=400)
            try:
                r = (QaRecord.objects
                     .select_related('feedback')
                     .get(id=qa_id))
            except QaRecord.DoesNotExist:
                return Response({'detail': 'QA 记录不存在'}, status=404)
            # OneToOne LEFT JOIN：无反馈时 r.feedback 为 None
            rating = r.feedback.rating if (hasattr(r, 'feedback') and r.feedback) else 0
            return Response({
                'row': {
                    'id': r.id,
                    'question': r.question,
                    'answer': r.answer,
                    'answer_type': r.answer_type,
                    'root_type': r.root_type,
                    'rating': rating,
                    'latency_total_ms': r.latency_total_ms,
                    'tokens_prompt': r.tokens_prompt,
                    'tokens_completion': r.tokens_completion,
                    'cost_estimate': float(r.cost_estimate),
                    'is_hit_cache': r.is_hit_cache,
                    'created_at': r.created_at.isoformat(),
                }
            })

        # —— 列表查询 ——
        qs = QaRecord.objects.all().select_related("feedback").order_by("-created_at")

        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        root_type = request.query_params.get("root_type")

        # 日期过滤：统一使用 __date__lte / __date__gte 而非 naive datetime，
        # 避免 created_at (timezone-aware) 与 naive datetime 比较触发 RuntimeWarning，
        # 同时保证 end_date 当天的所有记录都能被包含（不会被 end_date 00:00:00 截断）
        if start_date:
            try:
                start = datetime.fromisoformat(start_date).date()
                qs = qs.filter(created_at__date__gte=start)
            except ValueError:
                return Response({'detail': 'start_date 格式应为 YYYY-MM-DD'}, status=400)

        if end_date:
            try:
                end = datetime.fromisoformat(end_date).date()
                qs = qs.filter(created_at__date__lte=end)
            except ValueError:
                return Response({'detail': 'end_date 格式应为 YYYY-MM-DD'}, status=400)

        if root_type:
            qs = qs.filter(root_type=root_type)

        try:
            page = int(request.query_params.get("page") or 1)
            size = int(request.query_params.get("page_size") or 20)
        except (ValueError, TypeError):
            return Response({"detail": "page 和 page_size 必须为整数"}, status=400)

        if page < 1:
            page = 1
        if size < 1 or size > 100:
            size = 20
        offset = (page - 1) * size

        total = qs.count()
        rows = []
        for r in qs[offset:offset + size]:
            # OneToOne LEFT JOIN：无反馈时 r.feedback 为 None
            rating = r.feedback.rating if (hasattr(r, 'feedback') and r.feedback) else 0
            rows.append({
                "id": r.id,
                "question": r.question,
                "answer": r.answer,
                "answer_type": r.answer_type,
                "root_type": r.root_type,
                "rating": rating,
                "latency_total_ms": r.latency_total_ms,
                "tokens_prompt": r.tokens_prompt,
                "tokens_completion": r.tokens_completion,
                "cost_estimate": float(r.cost_estimate),
                "is_hit_cache": r.is_hit_cache,
                "created_at": r.created_at.isoformat(),
            })

        return Response({
            "total": total,
            "page": page,
            "page_size": size,
            "rows": rows,
        })


class BadFeedbackDetailView(APIView):
    """PUT /api/v1/analytics/bad-feedbacks/{id}/ 标记反馈状态

    - 修改反馈状态（pending→resolved/ignored）直接影响用户体验
    - 仅限管理员操作，需记录审计日志
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.write'

    def put(self, request, fb_id):
        status = request.data.get("status", "resolved")
        if status not in ("pending", "processing", "resolved", "ignored"):
            return Response({"detail": "无效的状态值"}, status=400)

        try:
            fb = QaFeedback.objects.get(id=fb_id)
            fb.status = status
            fb.save(update_fields=["status"])
            logger.info(
                f"feedback_status_updated fb_id={fb.id} status={status} user={request.user.username}"
            )
            return Response({"id": fb.id, "status": fb.status})
        except QaFeedback.DoesNotExist:
            return Response({"detail": "反馈不存在"}, status=404)


# ============================================================================
# 以下为系统监控类 API（需要 analytics 权限）
# ============================================================================

class SystemMetricsReportView(APIView):
    """GET /api/v1/analytics/system-metrics/?date=2026-07-24

    - 读取预计算的 SystemMetricsReport，避免实时聚合
    - date 参数可选，默认昨天
    - 仅管理员可访问（系统级敏感指标）
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request):
        date_str = request.query_params.get('date')
        if date_str:
            try:
                report_date = datetime.fromisoformat(date_str).date()
            except ValueError:
                return Response({'detail': '日期格式应为 YYYY-MM-DD'}, status=400)
        else:
            report_date = (timezone.now() - timedelta(days=1)).date()

        try:
            report = SystemMetricsReport.objects.get(report_date=report_date)
        except SystemMetricsReport.DoesNotExist:
            return Response({
                'date': str(report_date),
                'available': False,
                'message': '该日期的报表尚未生成，请等待凌晨聚合任务完成',
            })

        return Response({
            'date': str(report_date),
            'available': True,
            'total_qa': report.total_qa,
            'cache_hit_count': report.cache_hit_count,
            'normal_qa_count': report.normal_qa_count,
            # 正常请求延迟
            'p50_latency_total': report.p50_latency_total,
            'p95_latency_total': report.p95_latency_total,
            'p99_latency_total': report.p99_latency_total,
            'p50_latency_llm': report.p50_latency_llm,
            'p95_latency_llm': report.p95_latency_llm,
            'p50_latency_retrieval': report.p50_latency_retrieval,
            'p95_latency_retrieval': report.p95_latency_retrieval,
            'p50_ttfb': report.p50_ttfb,
            'p95_ttfb': report.p95_ttfb,
            # 缓存命中延迟
            'cache_hit_p50_latency': report.cache_hit_p50_latency,
            'cache_hit_p95_latency': report.cache_hit_p95_latency,
            # 比率
            'cache_hit_rate': report.cache_hit_rate,
            'llm_success_rate': report.llm_success_rate,
            'llm_timeout_rate': report.llm_timeout_rate,
            'embedding_error_rate': report.embedding_error_rate,
            'avg_tokens_per_second': report.avg_tokens_per_second,
            # Token & 成本
            'total_tokens_prompt': report.total_tokens_prompt,
            'total_tokens_completion': report.total_tokens_completion,
            'total_cost': float(report.total_cost),
            # 分布
            'latency_histogram': report.latency_histogram,
            'error_distribution': report.error_distribution,
        })


class OrgUsageReportView(APIView):
    """GET /api/v1/analytics/org-usage/?date=&department_id=&team_id=

    - 支持按日期、部门、团队筛选
    - department_id/team_id 可选，不传则返回所有
    - 同时返回部门汇总（team_id=-1）和团队明细
    - 权限控制：super_admin 看全部，部门负责人看本部门所有团队，
      团队负责人仅能看本团队（不能看其他团队或部门级汇总）
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.org.read'

    def get(self, request):
        # 权限模型：user.department FK 直接获取部门，团队管辖通过 UserTeamScopeRel 判定，
        # 部门管辖通过 UserDeptScopeRel 判定
        from apps.users.models import (
            UserDeptScopeRel, UserTeamScopeRel, GrantStatus, Team,
        )

        date_str = request.query_params.get('date')
        department_id = request.query_params.get('department_id')
        team_id = request.query_params.get('team_id')

        # --- 参数校验 ---
        try:
            if date_str:
                report_date = datetime.fromisoformat(date_str).date()
            else:
                report_date = (timezone.now() - timedelta(days=1)).date()
        except ValueError:
            return Response({'detail': '日期格式应为 YYYY-MM-DD'}, status=400)

        try:
            if department_id:
                department_id = int(department_id)
        except (ValueError, TypeError):
            return Response({'detail': 'department_id 必须为整数'}, status=400)

        try:
            if team_id:
                team_id = int(team_id)
        except (ValueError, TypeError):
            return Response({'detail': 'team_id 必须为整数'}, status=400)

        qs = OrgUsageReport.objects.filter(report_date=report_date)

        # --- 权限过滤：按用户角色分级控制可见范围 ---
        # super_admin 看全部；部门管理者看本部门所有团队 + 部门级汇总；
        # 团队负责人仅看本团队（不能看其他团队或部门级汇总）
        user = request.user
        if not user.is_super_admin:
            # User 直接有 department FK（单主部门）
            user_dept_id = user.department_id

            if not user_dept_id:
                # 无部门归属的非 admin 用户无法查看任何组织数据
                return Response({'detail': '无权限访问，请联系管理员'}, status=403)

            # 1) 跨部门访问拦截
            if department_id is not None and department_id != user_dept_id:
                return Response({'detail': '无权查看其他部门数据'}, status=403)

            # 2) 限制在本部门
            qs = qs.filter(department_id=user_dept_id)
            department_id = user_dept_id  # 固定为本部门，后续逻辑复用

            # 3) 判断部门管理者：有 UserDeptScopeRel 活跃授权 → 可见本部门所有团队 + 部门级汇总
            #    基于 RBAC 授权关系判定，非硬编码标记
            is_dept_manager = UserDeptScopeRel.objects.filter(
                user=user, dept_id=user_dept_id, status=GrantStatus.ACTIVE,
            ).exists()

            # 4) 团队级限制：通过 UserTeamScopeRel 获取用户管辖的团队（限本部门内）
            #    user.team 为单团队 FK，团队管辖授权通过 UserTeamScopeRel 显式绑定
            leading_team_ids = list(UserTeamScopeRel.objects.filter(
                user=user, status=GrantStatus.ACTIVE,
                team__department_id=user_dept_id,
            ).values_list('team_id', flat=True))

            if not is_dept_manager and leading_team_ids:
                # 团队负责人：仅允许访问 leading_team_ids 中的团队
                # 禁止访问 team_id=-1（部门级汇总），禁止访问其他团队
                if team_id is not None:
                    if team_id == -1:
                        return Response({'detail': '无权查看部门级汇总'}, status=403)
                    if team_id not in leading_team_ids:
                        return Response({'detail': '无权查看其他团队数据'}, status=403)
                else:
                    # 未指定 team_id，限制在用户负责的团队范围（过滤掉 -1 汇总和其他团队）
                    qs = qs.filter(team_id__in=leading_team_ids)
            elif not is_dept_manager:
                # 非部门管理者且非任何团队负责人 → 仅看个人维度不开放组织数据
                return Response({'detail': '无权限访问组织报表，请联系管理员'}, status=403)
            # else: 部门管理者 → 保持在本部门范围，允许查看所有团队 + 部门级汇总

        if department_id is not None:
            qs = qs.filter(department_id=department_id)
        if team_id is not None:
            # 前端传 -1 表示查看部门级汇总（team_id=-1 哨兵值）
            if team_id == -1:
                qs = qs.filter(team_id=-1)
            else:
                qs = qs.filter(team_id=team_id)
        else:
            # team_id 未指定时：默认仅返回团队级明细（team_id != -1），
            # 避免同时返回部门汇总+团队明细造成前端重复统计
            qs = qs.exclude(team_id=-1)

        rows = []
        for r in qs.order_by('department_id', 'team_id'):
            rows.append({
                'id': r.id,
                'report_date': str(r.report_date),
                'department_id': r.department_id,
                'department_name': r.department_name,
                'team_id': r.team_id,
                'team_name': r.team_name,
                'qa_count': r.qa_count,
                'user_count': r.user_count,
                'total_tokens': r.total_tokens,
                'total_cost': float(r.total_cost),
                'avg_latency_ms': r.avg_latency_ms,
                'p95_latency_ms': r.p95_latency_ms,
                'good_feedback_rate': r.good_feedback_rate,
                'cache_hit_count': r.cache_hit_count,
                'cache_hit_rate': r.cache_hit_rate,
            })

        return Response({
            'date': str(report_date),
            'count': len(rows),
            'rows': rows,
        })


class QueueDepthView(APIView):
    """GET /api/v1/analytics/queue-depth/?hours=24

    - 获取最近 N 小时的队列深度历史（PG 数据）
    - 同时返回当前实时深度（Redis）
    - 需具备 analytics:system:read 权限
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request):
        try:
            hours = int(request.query_params.get('hours', 24))
        except (ValueError, TypeError):
            return Response({'detail': 'hours 必须为整数'}, status=400)

        if hours < 1 or hours > 720:  # 限制 1h ~ 30 天
            return Response({'detail': 'hours 范围: 1-720'}, status=400)

        # --- 历史数据（PG）---
        from apps.analytics.utils import get_queue_depth_history
        history = get_queue_depth_history(hours=hours)

        # --- 当前实时深度（Redis）---
        try:
            from apps.analytics.realtime import get_queue_depth_snapshot
            current = get_queue_depth_snapshot()
        except Exception:
            logger.exception('[QueueDepth] Failed to get current snapshot')
            current = {}

        return Response({
            'hours': hours,
            'current': current,
            'history': history,
        })


class RealtimeSnapshotView(APIView):
    """GET /api/v1/analytics/realtime/

    - 获取今日实时指标快照（Redis 数据）
    - 用于 Dashboard 实时展示，5 分钟更新一次
    - 需具备 analytics:system:read 权限
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request):
        try:
            from apps.analytics.realtime import get_realtime_snapshot
            snapshot = get_realtime_snapshot()
            return Response(snapshot)
        except Exception:
            logger.exception('[RealtimeSnapshot] Failed to get snapshot')
            return Response({
                'date': timezone.now().date().isoformat(),
                'total_qa': 0,
                'cache_hits': 0,
                'llm_errors': 0,
                'error': 'snapshot_unavailable',
            }, status=200)


# ============================================================================
# 黄金测试集管理 Views
# ============================================================================

class GoldenDatasetListView(APIView):
    """GET/POST /api/v1/analytics/golden-datasets/"""
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get_permissions(self):
        # POST 为写操作,需要 write 权限;GET 只需 read
        # 在 get_permissions 阶段按 HTTP 方法切换 required_perm,避免读权限用户越权创建测试集
        if self.request.method == 'POST':
            self.required_perm = 'analytics.system.write'
        else:
            self.required_perm = 'analytics.system.read'
        return super().get_permissions()

    def get(self, request):
        from apps.analytics.models import GoldenDataset
        status = request.query_params.get('status')
        dataset_type = request.query_params.get('dataset_type')
        qs = GoldenDataset.objects.all().order_by('-updated_at')
        if status:
            qs = qs.filter(status=status)
        # 支持按 dataset_type 筛选(custom / regression_low_score),前端低分回归 Tab 用
        if dataset_type:
            qs = qs.filter(dataset_type=dataset_type)
        rows = list(qs[:100].values(
            'id', 'name', 'description', 'root_type', 'status',
            'dataset_type', 'question_count', 'version',
            'created_at', 'updated_at',
        ))
        # 补充 dataset_type 的中文展示名(避免前端维护映射表)
        type_label_map = dict(GoldenDataset.DATASET_TYPE_CHOICES)
        for r in rows:
            r['dataset_type_label'] = type_label_map.get(r['dataset_type'], r['dataset_type'])
        return Response({'rows': rows, 'count': len(rows)})

    def post(self, request):
        from apps.analytics.offline_eval import create_golden_dataset
        name = (request.data.get('name') or '').strip()
        if not name:
            return Response({'detail': 'name 必填'}, status=400)
        root_type = (request.data.get('root_type') or 'company_doc').strip()
        description = request.data.get('description', '')
        version = request.data.get('version', 'v1')
        ds = create_golden_dataset(
            name=name, root_type=root_type,
            description=description, version=version,
            created_by_id=request.user.id,
        )
        return Response({
            'id': ds.id, 'name': ds.name, 'root_type': ds.root_type,
            'status': ds.status, 'version': ds.version,
        })


class GoldenDatasetDetailView(APIView):
    """GET/PUT/DELETE /api/v1/analytics/golden-datasets/<id>/"""
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.write'

    def get_permissions(self):
        # GET 为读操作,只需 read 权限;PUT/DELETE 为写操作,需要 write 权限
        # 避免原实现中 GET 也要求 write 权限,导致只有读权限的用户无法查看单个测试集
        if self.request.method == 'GET':
            self.required_perm = 'analytics.system.read'
        else:
            self.required_perm = 'analytics.system.write'
        return super().get_permissions()

    def get(self, request, ds_id):
        from apps.analytics.models import GoldenDataset, GoldenQuestion
        from apps.analytics.serializers import GoldenQuestionSerializer
        try:
            ds = GoldenDataset.objects.get(id=ds_id)
        except GoldenDataset.DoesNotExist:
            return Response({'detail': '测试集不存在'}, status=404)
        # 一次查询同时拿到 relevant_doc_count（annotate Count）和 reference_answer（select_related）
        # 避免原实现中循环 q_obj.relevant_docs.count() 触发 N+1 COUNT 查询
        questions = (
            GoldenQuestion.objects
            .filter(dataset=ds)
            .order_by('order')
            .select_related('reference_answer')
            .annotate(relevant_doc_count=models.Count('relevant_docs'))
        )
        # 用 Serializer 替代手动循环构造 dict，字段集中管理且 relevant_doc_count/has_reference
        # 等计算字段已在序列化器中声明，便于其他接口复用
        questions_data = GoldenQuestionSerializer(questions, many=True).data
        return Response({
            'id': ds.id, 'name': ds.name, 'description': ds.description,
            'root_type': ds.root_type, 'status': ds.status,
            'dataset_type': ds.dataset_type,
            'dataset_type_label': ds.get_dataset_type_display(),
            'question_count': ds.question_count, 'version': ds.version,
            'questions': questions_data,
            # 建议移除阈值:前端据此标记"建议人工 review 移除",避免硬编码不一致
            'suggest_remove_passes': AnalyticsConfig.low_score_regression_suggest_remove_passes(),
        })

    def put(self, request, ds_id):
        from apps.analytics.models import GoldenDataset
        try:
            ds = GoldenDataset.objects.get(id=ds_id)
        except GoldenDataset.DoesNotExist:
            return Response({'detail': '测试集不存在'}, status=404)
        # 只允许更新安全字段，防止注入或类型异常
        allowed_fields = {'name', 'description', 'status', 'version'}
        for field in allowed_fields:
            if field in request.data:
                value = request.data[field]
                # name/version/description 转为字符串，status 校验是否合法
                if field == 'status' and value not in ('draft', 'active', 'archived'):
                    return Response({'detail': f'status 必须为 draft/active/archived'}, status=400)
                setattr(ds, field, str(value) if field != 'status' else value)
        ds.save()
        return Response({'id': ds.id, 'status': ds.status, 'name': ds.name})

    def delete(self, request, ds_id):
        from apps.analytics.models import GoldenDataset
        try:
            ds = GoldenDataset.objects.get(id=ds_id)
        except GoldenDataset.DoesNotExist:
            return Response({'detail': '测试集不存在'}, status=404)
        ds.delete()
        return Response({'ok': True})


class GoldenDatasetImportView(APIView):
    """POST /api/v1/analytics/golden-datasets/<id>/import/"""
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.write'

    def post(self, request, ds_id):
        from apps.analytics.offline_eval import import_questions_from_json
        questions_data = request.data.get('questions', [])
        if not questions_data:
            return Response({'detail': 'questions 必填'}, status=400)
        result = import_questions_from_json(
            dataset_id=ds_id,
            questions_data=questions_data,
            created_by_id=request.user.id,
        )
        return Response(result)


class GoldenDatasetExportView(APIView):
    """GET /api/v1/analytics/golden-datasets/<id>/export/"""
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request, ds_id):
        from apps.analytics.offline_eval import export_dataset_to_json
        data = export_dataset_to_json(ds_id)
        return Response({'dataset_id': ds_id, 'questions': data})


class GoldenQuestionView(APIView):
    """POST/DELETE /api/v1/analytics/golden-datasets/<ds_id>/questions/"""
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.write'

    def post(self, request, ds_id):
        from apps.analytics.offline_eval import import_questions_from_json
        q_data = {
            'question': request.data.get('question', ''),
            'question_type': request.data.get('question_type', 'factoid'),
            'difficulty': request.data.get('difficulty', 'medium'),
            'tags': request.data.get('tags', []),
            'relevant_doc_ids': request.data.get('relevant_doc_ids', []),
            'reference_answer': request.data.get('reference_answer', ''),
            'key_points': request.data.get('key_points', []),
        }
        result = import_questions_from_json(ds_id, [q_data], request.user.id)
        return Response(result)

    def delete(self, request, ds_id):
        from apps.analytics.models import GoldenQuestion
        question_id = request.query_params.get('question_id')
        if not question_id:
            return Response({'detail': 'question_id 必填'}, status=400)
        try:
            question_id = int(question_id)
        except (ValueError, TypeError):
            return Response({'detail': 'question_id 必须为整数'}, status=400)
        try:
            q = GoldenQuestion.objects.get(id=question_id, dataset_id=ds_id)
            q.delete()
            return Response({'ok': True})
        except GoldenQuestion.DoesNotExist:
            return Response({'detail': '问题不存在'}, status=404)


# ============================================================================
# 低分回归测试集 Views
# ============================================================================

class SiphonRegressionView(APIView):
    """POST /api/v1/analytics/regression/siphon/ - 手动触发低分沉淀

    从生产低分对话中取 top N 沉淀到回归测试集。同步执行(DB 操作,通常 1~2s),
    直接返回沉淀结果,前端刷新测试集列表查看新增问题。
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.write'

    def post(self, request):
        # 手动触发不受 LOW_SCORE_REGRESSION_ENABLED 开关限制
        # (开关只控制定时任务;管理员主动操作应生效)
        top_n = request.data.get('top_n')
        kwargs = {}
        if top_n:
            try:
                top_n = int(top_n)
                if top_n > 0:
                    kwargs['top_n'] = top_n
            except (ValueError, TypeError):
                return Response({'detail': 'top_n 必须为正整数'}, status=400)
        # 同步执行沉淀(DB 操作,通常 1~2s,直接返回结果避免前端轮询)
        from apps.analytics.regression_eval import siphon_low_score_qa_to_regression_set
        try:
            result = siphon_low_score_qa_to_regression_set(**kwargs)
            return Response({'ok': True, **result})
        except Exception as e:
            logger.exception('Siphon regression failed')
            return Response({'detail': f'沉淀失败: {e}'}, status=500)


class RunRegressionEvalView(APIView):
    """POST /api/v1/analytics/regression/eval/ - 手动触发低分回归全链路评估

    对低分回归测试集执行 检索→生成→12 维评估,更新 pass_count。
    成本较高(每问题 90~180s),异步派发 Celery 任务,前端提示后刷新查看结果。

    可选参数:
    - dataset_id: 指定测试集;不传则评估所有 regression_low_score 测试集
    - limit: 每个测试集最多评估的问题数(控制单次成本)
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.write'

    def post(self, request):
        from apps.analytics.tasks import run_regression_evaluation_task
        dataset_id = request.data.get('dataset_id')
        limit = request.data.get('limit')

        kwargs = {}
        if dataset_id:
            try:
                kwargs['dataset_id'] = int(dataset_id)
            except (ValueError, TypeError):
                return Response({'detail': 'dataset_id 必须为整数'}, status=400)
        if limit:
            try:
                kwargs['limit'] = int(limit)
            except (ValueError, TypeError):
                return Response({'detail': 'limit 必须为整数'}, status=400)

        # 异步派发:全链路评估耗时取决于问题数,200 条可能 30+ 分钟
        run_regression_evaluation_task.delay(**kwargs)
        return Response({
            'ok': True, 'queued': True,
            'message': '评估已派发,全链路评估耗时较长,请稍后刷新查看 pass_count 变化',
        })


# ============================================================================
# 离线评估执行 Views
# ============================================================================

class RunRetrievalEvalView(APIView):
    """POST /api/v1/analytics/eval/retrieval/ - 执行离线检索评估"""
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.write'

    def post(self, request):
        from apps.analytics.offline_eval import run_retrieval_evaluation
        dataset_id = request.data.get('dataset_id')
        if not dataset_id:
            return Response({'detail': 'dataset_id 必填'}, status=400)
        try:
            dataset_id = int(dataset_id)
        except (ValueError, TypeError):
            return Response({'detail': 'dataset_id 必须为整数'}, status=400)
        try:
            report = run_retrieval_evaluation(
                dataset_id=dataset_id,
                user=request.user,
            )
            return Response({
                'ok': True,
                'report_id': report.id,
                'recall_at_5': report.recall_at_5,
                'recall_at_10': report.recall_at_10,
                'recall_at_20': report.recall_at_20,
                'mrr': report.mrr,
                'ndcg_at_10': report.ndcg_at_10,
                'questions_with_hits': report.questions_with_hits,
                'questions_without_hits': report.questions_without_hits,
            })
        except Exception as e:
            logger.exception('Retrieval eval failed')
            return Response({'detail': f'评估失败: {e}'}, status=500)


class RunAnswerEvalView(APIView):
    """POST /api/v1/analytics/eval/answer/ - 执行离线回答质量评估"""
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.write'

    def post(self, request):
        from apps.analytics.offline_eval import run_answer_quality_evaluation
        dataset_id = request.data.get('dataset_id')
        if not dataset_id:
            return Response({'detail': 'dataset_id 必填'}, status=400)
        try:
            dataset_id = int(dataset_id)
        except (ValueError, TypeError):
            return Response({'detail': 'dataset_id 必须为整数'}, status=400)
        try:
            max_questions = int(request.data.get('max_questions', 50))
        except (ValueError, TypeError):
            return Response({'detail': 'max_questions 必须为整数'}, status=400)
        # 限制评估数量，防止 LLM 成本失控
        max_questions = max(1, min(max_questions, 100))
        try:
            results = run_answer_quality_evaluation(
                dataset_id=int(dataset_id),
                user=request.user,
                max_questions=max_questions,
            )
            return Response({
                'ok': True,
                'evaluated_count': len(results),
                'results': results[:20],
            })
        except Exception as e:
            logger.exception('Answer eval failed')
            return Response({'detail': f'评估失败: {e}'}, status=500)


class RetrievalReportListView(APIView):
    """GET /api/v1/analytics/eval/retrieval-reports/"""
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request):
        from apps.analytics.models import RetrievalQualityReport
        qs = RetrievalQualityReport.objects.select_related('dataset').order_by('-created_at')
        rows = list(qs[:50].values(
            'id', 'dataset_id', 'eval_batch_id',
            'recall_at_5', 'recall_at_10', 'recall_at_20', 'mrr', 'ndcg_at_5', 'ndcg_at_10',
            'vector_recall_at_10', 'bm25_recall_at_10', 'hybrid_recall_at_10', 'rerank_recall_at_10',
            'total_questions', 'questions_with_hits', 'questions_without_hits',
            'status', 'created_at',
        ))
        return Response({'rows': rows, 'count': len(rows)})


# ============================================================================
# 文档质量 Views
# ============================================================================

class DocumentQualityReportView(APIView):
    """GET /api/v1/analytics/doc-quality/?start_date=&end_date=&dept_id=&team_id= - 文档质量汇总"""
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request):
        from apps.analytics.doc_quality import get_document_quality_summary
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        root_type = request.query_params.get('root_type')
        dept_id, team_id = _parse_org_scope(request)
        if start_date:
            try:
                datetime.fromisoformat(start_date)
            except ValueError:
                return Response({'detail': 'start_date 格式应为 YYYY-MM-DD'}, status=400)
        if end_date:
            try:
                datetime.fromisoformat(end_date)
            except ValueError:
                return Response({'detail': 'end_date 格式应为 YYYY-MM-DD'}, status=400)
        return Response({
            **get_document_quality_summary(
                start_date, end_date, root_type,
                dept_id=dept_id, team_id=team_id,
            ),
            'dept_id': dept_id,
            'team_id': team_id,
        })


class RunDocQualityEvalView(APIView):
    """POST /api/v1/analytics/doc-quality/evaluate/ - 触发文档质量评估"""
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.write'

    def post(self, request):
        from apps.analytics.doc_quality import evaluate_document_quality, batch_evaluate_document_quality
        doc_id = request.data.get('document_id')
        if doc_id:
            try:
                doc_id = int(doc_id)
            except (ValueError, TypeError):
                return Response({'detail': 'document_id 必须为整数'}, status=400)
            report = evaluate_document_quality(doc_id)
            return Response({'ok': True, 'report_id': report.id, 'score': report.quality_score})
        try:
            days = int(request.data.get('days', 7))
        except (ValueError, TypeError):
            return Response({'detail': 'days 必须为整数'}, status=400)
        days = max(1, min(days, 30))
        summary = batch_evaluate_document_quality(days=days)
        return Response({'ok': True, 'summary': summary})


class DocumentQualityReportListView(APIView):
    """GET /api/v1/analytics/doc-quality/reports/?dept_id=&team_id="""
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request):
        from apps.analytics.models import DocumentQualityReport
        from apps.analytics.serializers import DocumentQualityReportSerializer
        root_type = request.query_params.get('root_type')
        min_score = request.query_params.get('min_score')
        dept_id, team_id = _parse_org_scope(request)

        qs = DocumentQualityReport.objects.select_related('document').order_by('-created_at')
        if root_type:
            qs = qs.filter(document__root_type=root_type)
        qs = _apply_org_filter_on_doc(qs, dept_id, team_id, doc_prefix='document__')
        if min_score:
            try:
                qs = qs.filter(quality_score__gte=float(min_score))
            except (ValueError, TypeError):
                return Response({'detail': 'min_score 必须为数字'}, status=400)

        total = qs.count()
        rows = DocumentQualityReportSerializer(qs[:50], many=True).data
        return Response({'total': total, 'rows': rows, 'dept_id': dept_id, 'team_id': team_id})


# ============================================================================
# 多维度评估 Views
# ============================================================================

class MultiDimensionScoreView(APIView):
    """GET /api/v1/analytics/multi-dim-scores/ - 多维度评估结果"""
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request):
        from apps.analytics.models import MultiDimensionScore
        qa_id = request.query_params.get('qa_record_id')
        dimension = request.query_params.get('dimension')
        start_date = request.query_params.get('start_date')

        qs = MultiDimensionScore.objects.select_related('qa_record').order_by('-created_at')
        if qa_id:
            try:
                qs = qs.filter(qa_record_id=int(qa_id))
            except (ValueError, TypeError):
                return Response({'detail': 'qa_record_id 必须为整数'}, status=400)
        if dimension:
            qs = qs.filter(dimension=dimension)
        if start_date:
            try:
                start_dt = datetime.fromisoformat(start_date).date()
                qs = qs.filter(created_at__date__gte=start_dt)
            except ValueError:
                return Response({'detail': 'start_date 格式应为 YYYY-MM-DD'}, status=400)

        # 汇总：用一次 GROUP BY 替代原 N 次逐维度查询，避免 N+1
        dim_agg_qs = qs.values('dimension').annotate(
            count=models.Count('id'),
            avg_score=models.Avg('score'),
            min_score=models.Min('score'),
            max_score=models.Max('score'),
        )
        dim_agg = {}
        for r in dim_agg_qs:
            dim_agg[r['dimension']] = {
                'count': r['count'],
                'avg_score': round(r['avg_score'] or 0, 4),
                'min_score': round(r['min_score'] or 0, 4),
                'max_score': round(r['max_score'] or 0, 4),
            }

        total = qs.count()
        rows = list(qs[:50].values(
            'id', 'qa_record_id', 'dimension', 'score', 'reason',
            'eval_model', 'eval_cost', 'status', 'created_at',
        ))
        return Response({'total': total, 'dimension_summary': dim_agg, 'rows': rows})


class RunMultiDimEvalView(APIView):
    """POST /api/v1/analytics/multi-dim-eval/ - 手动触发单条 QA 的 DeepEval 12 维评估

    用于看板中"手动评估指定 QA"场景:运营发现某条对话异常,手动触发评估。
    实际启用维度由 EVAL_DISPLAY_DIMENSIONS 控制(默认 12 维,评估=展示强绑定)。

    异步执行:POST 立即派发 Celery 任务并返回 eval_batch_id,前端通过轮询
    qa-detail 接口检查该 batch_id 的评估结果是否落库。
    改为异步的原因:12 维 LLM 评估串行耗时 90~180s+,同步 HTTP 请求易被
    网关/浏览器超时断开,且会阻塞 Django dev server 其他请求。
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.write'

    def post(self, request):
        from apps.analytics.production_eval import evaluate_sampled_qa

        qa_id = request.data.get('qa_record_id')
        if not qa_id:
            return Response({'detail': 'qa_record_id 必填'}, status=400)
        try:
            qa_id = int(qa_id)
        except (ValueError, TypeError):
            return Response({'detail': 'qa_record_id 必须为整数'}, status=400)

        try:
            qa = QaRecord.objects.get(id=qa_id)
        except QaRecord.DoesNotExist:
            return Response({'detail': 'QA记录不存在'}, status=404)

        # 预检:无检索上下文的 QA 无法评估,提前返回避免无意义排队
        from apps.analytics.production_eval import _build_context_list
        contexts = _build_context_list(qa)
        if not contexts:
            return Response({'detail': '该 QA 无检索上下文,无法评估'}, status=400)

        # 预生成 eval_batch_id 作为本次评估的唯一标识
        # 前端通过该 id 在 qa-detail 接口轮询本次评估落库的维度数
        eval_batch_id = f'manual_{timezone.now().strftime("%Y%m%d%H%M%S")}_{qa_id}'

        # 派发 Celery 任务:手动评估场景跳过日预算检查(不计入生产配额)
        evaluate_sampled_qa.delay(
            qa_id,
            skip_budget_check=True,
            eval_batch_id=eval_batch_id,
        )

        return Response({
            'ok': True,
            'queued': True,
            'qa_id': qa_id,
            'eval_batch_id': eval_batch_id,
            'message': '评估已派发,请通过 eval_batch_id 轮询结果',
        })


# ============================================================================
# 覆盖率 & 反馈闭环 Views
# ============================================================================

class CoverageReportView(APIView):
    """GET /api/v1/analytics/coverage/ - 覆盖率报告"""
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request):
        from apps.analytics.coverage import (
            analyze_hot_query_coverage, detect_knowledge_gaps,
            detect_duplicate_chunks, analyze_domain_coverage,
        )
        try:
            days = int(request.query_params.get('days', 7))
        except (ValueError, TypeError):
            return Response({'detail': 'days 必须为整数'}, status=400)
        # 限制 days 范围：1-30 天，防止过大范围扫描全表
        days = max(1, min(days, 30))

        coverage = analyze_hot_query_coverage(days)
        gaps = detect_knowledge_gaps(days)
        duplicates = detect_duplicate_chunks()
        domain = analyze_domain_coverage(days)

        return Response({
            'coverage': coverage,
            'gaps': gaps[:20],
            'gap_count': len(gaps),
            'duplicates': duplicates,
            'domain': domain,
        })


class FeedbackLoopView(APIView):
    """POST /api/v1/analytics/feedback-loop/ - 执行反馈闭环分析"""
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.write'

    def post(self, request):
        from apps.analytics.coverage import auto_link_feedback_to_chunks
        try:
            days = int(request.data.get('days', 7))
        except (ValueError, TypeError):
            return Response({'detail': 'days 必须为整数'}, status=400)
        days = max(1, min(days, 30))
        result = auto_link_feedback_to_chunks(days=days)
        return Response(result)


class GenerateCoverageReportView(APIView):
    """POST /api/v1/analytics/coverage/generate/ - 生成覆盖率报告"""
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.write'

    def post(self, request):
        from apps.analytics.coverage import generate_coverage_report
        try:
            days = int(request.data.get('days', 7))
        except (ValueError, TypeError):
            return Response({'detail': 'days 必须为整数'}, status=400)
        days = max(1, min(days, 30))
        report = generate_coverage_report(days=days)
        return Response({
            'ok': True,
            'report_id': report.id,
            'report_date': str(report.report_date),
            'coverage_rate': report.hot_query_coverage_rate,
            'gap_count': report.gap_count,
        })


class CoverageReportListView(APIView):
    """GET /api/v1/analytics/coverage/reports/ - 历史覆盖率报告列表

    - 返回最近 50 条报告记录，按日期倒序
    - 供前端「历史报告」面板展示，支持下载和删除
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request):
        from apps.analytics.models import CoverageReport
        # 最近 50 条，避免全表返回
        rows = list(
            CoverageReport.objects
            .order_by('-report_date', '-created_at')
            .values(
                'id', 'report_date', 'total_hot_queries', 'covered_queries',
                'uncovered_queries', 'hot_query_coverage_rate', 'gap_count',
                'duplicate_chunk_rate', 'duplicate_chunk_count',
                'feedback_loop_count', 'feedback_resolved_count',
                'created_at',
            )[:50]
        )
        return Response({'rows': rows, 'count': len(rows)})


class CoverageReportDetailView(APIView):
    """DELETE /api/v1/analytics/coverage/reports/<id>/ - 删除覆盖率报告

    - 仅允许删除历史报告，不影响当前覆盖率展示
    - 删除操作记录审计日志
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.write'

    def delete(self, request, report_id):
        from apps.analytics.models import CoverageReport
        try:
            report = CoverageReport.objects.get(id=report_id)
        except CoverageReport.DoesNotExist:
            return Response({'detail': '报告不存在'}, status=404)
        report.delete()
        logger.info(
            f'coverage_report_deleted report_id={report_id} date={report.report_date} user={request.user.username}'
        )
        return Response({'ok': True})


class CoverageReportExportView(APIView):
    """GET /api/v1/analytics/coverage/reports/<id>/export/ - 导出覆盖率报告为 Excel

    - 将单条报告的完整数据导出为 .xlsx 文件
    - 包含：覆盖率概览、知识空白详情、部门/团队覆盖明细
    - 使用 openpyxl 生成多 Sheet Excel
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request, report_id):
        from apps.analytics.models import CoverageReport
        try:
            report = CoverageReport.objects.get(id=report_id)
        except CoverageReport.DoesNotExist:
            return Response({'detail': '报告不存在'}, status=404)

        # 使用 openpyxl 生成 Excel，比 csv 支持多 Sheet + 格式化
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()

        # ===== Sheet 1: 覆盖率概览 =====
        ws1 = wb.active
        ws1.title = '覆盖率概览'
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')

        overview = [
            ('报告ID', report.id),
            ('报告日期', str(report.report_date)),
            ('生成时间', report.created_at.strftime('%Y-%m-%d %H:%M')),
            ('', ''),
            ('热门查询总数', report.total_hot_queries),
            ('已覆盖查询数', report.covered_queries),
            ('未覆盖查询数', report.uncovered_queries),
            ('热门问题覆盖率', f'{report.hot_query_coverage_rate * 100:.1f}%'),
            ('知识空白数', report.gap_count),
            ('重复切片率', f'{report.duplicate_chunk_rate * 100:.1f}%'),
            ('重复切片数', report.duplicate_chunk_count),
            ('反馈关联数', report.feedback_loop_count),
            ('反馈已解决数', report.feedback_resolved_count),
        ]
        for row_idx, (label, value) in enumerate(overview, 1):
            ws1.cell(row=row_idx, column=1, value=label).font = header_font
            ws1.cell(row=row_idx, column=1).fill = header_fill
            ws1.cell(row=row_idx, column=2, value=value)
        ws1.column_dimensions['A'].width = 20
        ws1.column_dimensions['B'].width = 30

        # ===== Sheet 2: 知识空白详情 =====
        ws2 = wb.create_sheet('知识空白')
        gaps = report.gap_queries or []
        ws2.append(['查询内容', '出现次数', '改进建议'])
        for col in range(1, 4):
            cell = ws2.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for g in gaps:
            ws2.append([
                g.get('query', ''),
                g.get('count', 0),
                g.get('suggestion', ''),
            ])
        ws2.column_dimensions['A'].width = 40
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 50

        # ===== Sheet 3: 部门/团队覆盖明细 =====
        ws3 = wb.create_sheet('部门覆盖')
        domain = report.domain_coverage or {}
        domain_list = domain.get('domain_coverage', []) if isinstance(domain, dict) else domain
        ws3.append(['部门/团队', '文档数', '切片数', '占比', '命中率', '下属团队数'])
        for col in range(1, 7):
            cell = ws3.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for d in domain_list:
            teams = d.get('teams', []) if isinstance(d, dict) else []
            # 兼容 dict 和 list 两种数据格式
            doc_count = d.get('doc_count', 0) if isinstance(d, dict) else 0
            chunk_count = d.get('chunk_count', 0) if isinstance(d, dict) else 0
            share = d.get('占比', 0) if isinstance(d, dict) else 0
            hit_rate = d.get('query_hit_rate', 0) if isinstance(d, dict) else 0
            name = d.get('name', '') if isinstance(d, dict) else str(d)
            ws3.append([
                name,
                doc_count,
                chunk_count,
                f'{share * 100:.1f}%' if isinstance(share, (int, float)) else str(share),
                f'{hit_rate * 100:.1f}%' if isinstance(hit_rate, (int, float)) else str(hit_rate),
                len(teams),
            ])
            # 如果有子团队，在下一行缩进展示
            if isinstance(teams, list):
                for team_item in teams:
                    if isinstance(team_item, (list, tuple)) and len(team_item) == 2:
                        team_name, team_data = team_item
                        t_docs = team_data.get('doc_count', 0) if isinstance(team_data, dict) else 0
                        t_chunks = team_data.get('chunk_count', 0) if isinstance(team_data, dict) else 0
                        ws3.append([f'  └ {team_name}', t_docs, t_chunks, '', '', ''])
        ws3.column_dimensions['A'].width = 25
        for col in 'BCDEF':
            ws3.column_dimensions[col].width = 12

        # 输出为 HTTP 响应，浏览器直接下载
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'coverage_report_{report.report_date}.xlsx'
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        logger.info(
            f'coverage_report_exported report_id={report_id} user={request.user.username}'
        )
        return response


# ============================================================================
# 评估看板 (DeepEval 12 维生产评估结果展示)
# ============================================================================
# 参考 LangSmith / Phoenix / Datadog 风格:
# - overview: 顶部 KPI(12 维均分按 4 大类分组 + 评估量 + 低分占比 + 安全告警)
# - trend: 时间序列(按天聚合,支持维度切换)
# - low-score-qa: 低分对话 Top N(按 QA 总分升序)
# - qa-detail: 单条 QA 完整明细(12 维 score+reason + 完整对话)

# 维度分组定义(与 deepeval_metrics.py 保持一致)
_DIMENSION_GROUPS = {
    'retrieval': ['context_relevancy'],
    'quality': ['faithfulness', 'hallucination', 'answer_relevancy',
                'completeness', 'conciseness', 'clarity'],
    'safety': ['toxicity', 'bias'],
    'business': ['professionalism', 'helpfulness', 'actionability'],
}


def _parse_dashboard_days(request) -> int:
    """解析看板 days 参数(默认 7,范围 1-90)"""
    try:
        days = int(request.query_params.get('days', 7))
    except (ValueError, TypeError):
        days = 7
    return max(1, min(days, 90))


class EvalDashboardOverviewView(APIView):
    """GET /api/v1/analytics/eval-dashboard/overview/?days=7&root_type=&dept_id=&team_id=

    看板顶部 KPI:
    - total_evaluated: 评估过的 QA 数(distinct qa_record_id)
    - total_qa: 时间窗口内 QA 总数(用于覆盖率计算)
    - low_score_count: 平均分 < threshold 的 QA 数
    - safety_alert_count: toxicity<0.5 或 bias<0.5 的 QA 数(安全告警)
    - dimension_groups: 4 大类均分 + 各维度均分(雷达图用)
    组织筛选(dept_id/team_id):按提问用户的部门/团队归属过滤,team 有值时忽略 dept。
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request):
        from apps.analytics.models import MultiDimensionScore
        days = _parse_dashboard_days(request)
        root_type = request.query_params.get('root_type', '').strip()
        dept_id, team_id = _parse_org_scope(request)
        threshold = 0.5

        since = timezone.now() - timedelta(days=days)

        # 基础 queryset(按时间 + 可选 root_type + 组织筛选)
        scores_qs = MultiDimensionScore.objects.filter(created_at__gte=since)
        qa_qs = QaRecord.objects.filter(created_at__gte=since, is_success=True).exclude(answer_type='refused')
        if root_type:
            scores_qs = scores_qs.filter(qa_record__root_type=root_type)
            qa_qs = qa_qs.filter(root_type=root_type)
        # 组织筛选(按 QaRecord.user 归属 JOIN):
        # scores_qs JOIN 通过 qa_record__ 关联 QaRecord,qa_qs 直接是 QaRecord
        scores_qs = _apply_org_filter_on_qa(scores_qs, dept_id, team_id, qa_prefix='qa_record__')
        qa_qs = _apply_org_filter_on_qa(qa_qs, dept_id, team_id)

        # 1. 评估量(去重 qa_record_id)
        total_evaluated = scores_qs.values('qa_record_id').distinct().count()
        total_qa = qa_qs.count()

        # 无评估数据时直接返回空结构,避免前端渲染 0 值雷达图/维度标签造成"写死数据"的误解
        if total_evaluated == 0:
            return Response({
                'days': days,
                'root_type': root_type or 'all',
                'total_evaluated': 0,
                'total_qa': total_qa,
                'coverage_rate': 0,
                'low_score_count': 0,
                'low_score_rate': 0,
                'safety_alert_count': 0,
                'threshold': threshold,
                'date_range': [],
                'dimension_groups': {},
                # 即使无评估数据也返回展示维度白名单,前端据此渲染维度组的空态结构
                'display_dimensions': AnalyticsConfig.eval_display_dimensions(),
            })

        # 2. 各维度均分(一次 GROUP BY 拿全)
        dim_agg = scores_qs.values('dimension').annotate(
            avg_score=models.Avg('score'),
            count=models.Count('id'),
        )
        dim_map = {r['dimension']: {'avg': float(r['avg_score'] or 0), 'count': r['count']} for r in dim_agg}

        # 3. 每个 QA 的均分(用于低分统计)
        # 用 GROUP BY qa_record_id + Avg,一次查询拿到所有 QA 的均分
        qa_avg = scores_qs.values('qa_record_id').annotate(
            avg_score=models.Avg('score'),
        )
        qa_avg_list = list(qa_avg)
        low_score_count = sum(1 for r in qa_avg_list if float(r['avg_score'] or 0) < threshold)

        # 4. 安全告警:toxicity 或 bias 维度 < 0.5 的 QA 数(精确查 toxicity/bias 维度)
        safety_qs = scores_qs.filter(
            dimension__in=['toxicity', 'bias'], score__lt=0.5
        ).values_list('qa_record_id', flat=True).distinct()
        safety_alert_count = len(set(safety_qs))

        # 5. 每个维度最近 7 天每日均分(用于前端 sparkline + 环比)
        # 一次 TruncDate+dimension 聚合,拿到所有 (date, dimension, avg) 三元组
        today = timezone.now().date()
        trend_start = today - timedelta(days=max(days, 7) - 1)
        trend_qs = scores_qs.filter(
            created_at__date__gte=trend_start,
        ).annotate(
            _date=models.functions.TruncDate('created_at'),
        ).values('_date', 'dimension').annotate(
            avg_score=models.Avg('score'),
        ).order_by('_date', 'dimension')
        # 按维度分桶: {dim: ['0.8','0.82',...]} 每天一个 float,缺失天用 0 填充
        dim_trend_map = {}
        dim_dates = []
        for r in trend_qs:
            d = r['dimension']
            date_str = r['_date'].isoformat()
            if date_str not in dim_dates:
                dim_dates.append(date_str)
            if d not in dim_trend_map:
                dim_trend_map[d] = {}
            dim_trend_map[d][date_str] = round(float(r['avg_score'] or 0), 4)
        # 补齐所有日期,确保每个维度 trend_7d 长度一致
        dim_dates.sort()
        for d in dim_trend_map:
            dim_trend_map[d] = [dim_trend_map[d].get(dt, 0) for dt in dim_dates]

        # 前一周期均分(环比用):前 7 天(不含当前窗口)的维度均分
        prev_since = since - timedelta(days=days)
        prev_end = since
        prev_scores_qs = MultiDimensionScore.objects.filter(
            created_at__gte=prev_since, created_at__lt=prev_end,
        )
        if root_type:
            prev_scores_qs = prev_scores_qs.filter(qa_record__root_type=root_type)
        # 环比同样按组织过滤,避免口径不一致
        prev_scores_qs = _apply_org_filter_on_qa(prev_scores_qs, dept_id, team_id, qa_prefix='qa_record__')
        prev_dim_agg = prev_scores_qs.values('dimension').annotate(
            prev_avg=models.Avg('score'),
        )
        prev_dim_map = {r['dimension']: round(float(r['prev_avg'] or 0), 4) for r in prev_dim_agg}

        # 6. 组装 4 大类
        groups = {}
        for group_name, dims in _DIMENSION_GROUPS.items():
            dim_list = []
            scores_in_group = []
            for d in dims:
                info = dim_map.get(d)
                if info:
                    cur_avg = round(info['avg'], 4)
                    prev_avg = prev_dim_map.get(d, 0)
                    # 环比变化率(prev 为 0 时无意义,置 None)
                    mom_change = round((cur_avg - prev_avg) / prev_avg, 4) if prev_avg > 0 else None
                    dim_list.append({
                        'name': d,
                        'avg': cur_avg,
                        'count': info['count'],
                        'prev_avg': prev_avg,
                        'mom_change': mom_change,
                        'trend_7d': dim_trend_map.get(d, []),
                    })
                    scores_in_group.append(info['avg'])
            groups[group_name] = {
                'avg_score': round(sum(scores_in_group) / len(scores_in_group), 4) if scores_in_group else 0,
                'dimensions': dim_list,
            }

        return Response({
            'days': days,
            'root_type': root_type or 'all',
            'dept_id': dept_id,
            'team_id': team_id,
            'total_evaluated': total_evaluated,
            'total_qa': total_qa,
            'coverage_rate': round(total_evaluated / total_qa, 4) if total_qa else 0,
            'low_score_count': low_score_count,
            'low_score_rate': round(low_score_count / total_evaluated, 4) if total_evaluated else 0,
            'safety_alert_count': safety_alert_count,
            'threshold': threshold,
            'date_range': dim_dates,
            'dimension_groups': groups,
            # 展示维度白名单：前端据此过滤「回答质量」页的维度画像，
            # 未在白名单中的维度不再展示（由 SystemConfig.EVAL_DISPLAY_DIMENSIONS 控制）
            'display_dimensions': AnalyticsConfig.eval_display_dimensions(),
        })


class EvalDashboardTrendView(APIView):
    """GET /api/v1/analytics/eval-dashboard/trend/?days=7&root_type=&dimension=&dept_id=&team_id=

    看板趋势线:按天聚合各维度均分。
    - dimension 为空:返回 12 维全部趋势(前端可切换显示)
    - dimension 指定:只返回该维度趋势
    组织筛选(dept_id/team_id):按提问用户归属过滤。
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request):
        from apps.analytics.models import MultiDimensionScore
        days = _parse_dashboard_days(request)
        root_type = request.query_params.get('root_type', '').strip()
        dimension = request.query_params.get('dimension', '').strip()
        dept_id, team_id = _parse_org_scope(request)

        since = timezone.now() - timedelta(days=days)
        qs = MultiDimensionScore.objects.filter(created_at__gte=since)
        if root_type:
            qs = qs.filter(qa_record__root_type=root_type)
        qs = _apply_org_filter_on_qa(qs, dept_id, team_id, qa_prefix='qa_record__')
        if dimension:
            qs = qs.filter(dimension=dimension)

        # 按天 + 维度聚合:TruncDate(created_at) + dimension + Avg(score)
        # 一次查询拿到所有 (date, dimension, avg) 三元组
        trend_qs = qs.annotate(
            date=models.functions.TruncDate('created_at'),
        ).values('date', 'dimension').annotate(
            avg_score=models.Avg('score'),
            count=models.Count('id'),
        ).order_by('date', 'dimension')

        # 组装成 {dates: [...], series: [{dimension, scores: [{date, avg, count}]}]}
        dates_set = sorted({r['date'].isoformat() for r in trend_qs})
        dim_series = {}
        for r in trend_qs:
            d = r['dimension']
            if d not in dim_series:
                dim_series[d] = []
            dim_series[d].append({
                'date': r['date'].isoformat(),
                'avg': round(float(r['avg_score'] or 0), 4),
                'count': r['count'],
            })

        return Response({
            'days': days,
            'root_type': root_type or 'all',
            'dept_id': dept_id,
            'team_id': team_id,
            'dimension': dimension or 'all',
            'dates': dates_set,
            'series': [
                {'dimension': d, 'scores': s}
                for d, s in dim_series.items()
            ],
        })


class EvalDashboardLowScoreView(APIView):
    """GET /api/v1/analytics/eval-dashboard/low-score-qa/?days=7&limit=20&threshold=0.5&root_type=&dept_id=&team_id=

    低分对话 Top N(按 QA 均分升序)。
    返回每个 QA 的:question/answer 摘要 + 均分 + 最低维度 + 最低分 + root_type + 时间。
    前端点击展开调 qa-detail 看完整明细。
    组织筛选(dept_id/team_id):按提问用户归属过滤。
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request):
        from apps.analytics.models import MultiDimensionScore
        days = _parse_dashboard_days(request)
        root_type = request.query_params.get('root_type', '').strip()
        dept_id, team_id = _parse_org_scope(request)
        try:
            limit = int(request.query_params.get('limit', 20))
        except (ValueError, TypeError):
            limit = 20
        limit = max(1, min(limit, 100))
        try:
            threshold = float(request.query_params.get('threshold', 0.5))
        except (ValueError, TypeError):
            threshold = 0.5

        since = timezone.now() - timedelta(days=days)
        qs = MultiDimensionScore.objects.filter(created_at__gte=since)
        if root_type:
            qs = qs.filter(qa_record__root_type=root_type)
        qs = _apply_org_filter_on_qa(qs, dept_id, team_id, qa_prefix='qa_record__')

        # 每个 QA 的均分 + 最低分维度(子查询取 min)
        qa_agg = qs.values('qa_record_id').annotate(
            avg_score=models.Avg('score'),
            min_score=models.Min('score'),
        ).filter(avg_score__lt=threshold).order_by('avg_score')[:limit]

        if not qa_agg:
            return Response({'total': 0, 'threshold': threshold, 'rows': [],
                             'dept_id': dept_id, 'team_id': team_id})

        # 批量取 QaRecord 详情(避免 N+1)
        qa_ids = [r['qa_record_id'] for r in qa_agg]
        qa_map = {
            q.id: q for q in QaRecord.objects.filter(id__in=qa_ids).only(
                'id', 'question', 'answer', 'root_type', 'created_at', 'user_id'
            )
        }

        # 批量取每个 QA 的最低分维度(一次 GROUP BY)
        min_dim_qs = qs.filter(qa_record_id__in=qa_ids).values(
            'qa_record_id', 'dimension', 'score'
        )
        # 对每个 qa_id 取 score 最低的 dimension
        qa_min_dim = {}
        for r in min_dim_qs:
            qid = r['qa_record_id']
            if qid not in qa_min_dim or r['score'] < qa_min_dim[qid]['score']:
                qa_min_dim[qid] = {'dimension': r['dimension'], 'score': float(r['score'])}

        rows = []
        for r in qa_agg:
            qid = r['qa_record_id']
            q = qa_map.get(qid)
            if not q:
                continue
            min_info = qa_min_dim.get(qid, {'dimension': '-', 'score': 0})
            rows.append({
                'qa_record_id': qid,
                'question': q.question[:80],
                'answer': q.answer[:120],
                'avg_score': round(float(r['avg_score'] or 0), 4),
                'min_dimension': min_info['dimension'],
                'min_score': round(min_info['score'], 4),
                'root_type': q.root_type,
                'created_at': q.created_at.isoformat(),
            })

        return Response({'total': len(rows), 'threshold': threshold, 'rows': rows,
                         'dept_id': dept_id, 'team_id': team_id})


class EvalDashboardQaDetailView(APIView):
    """GET /api/v1/analytics/eval-dashboard/qa-detail/?qa_record_id=123

    单条 QA 完整明细:
    - qa: 完整对话(question/answer/root_type/created_at/user/retrieval_hits)
    - scores: 12 维 score+reason+eval_model+latency(按 4 大类顺序)
    - avg_score: 总均分
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request):
        from apps.analytics.models import MultiDimensionScore
        qa_id = request.query_params.get('qa_record_id')
        if not qa_id:
            return Response({'detail': 'qa_record_id 必填'}, status=400)
        try:
            qa_id = int(qa_id)
        except (ValueError, TypeError):
            return Response({'detail': 'qa_record_id 必须为整数'}, status=400)

        try:
            qa = QaRecord.objects.select_related('user').get(id=qa_id)
        except QaRecord.DoesNotExist:
            return Response({'detail': 'QA 不存在'}, status=404)

        scores = list(
            MultiDimensionScore.objects.filter(qa_record_id=qa_id)
            .order_by('created_at')
            .values('dimension', 'score', 'reason', 'eval_model', 'eval_latency_ms', 'eval_batch_id', 'created_at')
        )

        # 按 4 大类顺序排序(未分类的维度放最后)
        dim_order = {d: (g, i) for g, (group, dims) in enumerate(_DIMENSION_GROUPS.items()) for i, d in enumerate(dims)}
        scores.sort(key=lambda s: dim_order.get(s['dimension'], (99, 99)))

        avg_score = sum(float(s['score'] or 0) for s in scores) / max(len(scores), 1)

        return Response({
            'qa': {
                'id': qa.id,
                'question': qa.question,
                'answer': qa.answer,
                'root_type': qa.root_type,
                'answer_type': qa.answer_type,
                'created_at': qa.created_at.isoformat(),
                'user': qa.user.username if qa.user else '-',
                'retrieval_hits': qa.retrieval_hits[:10],
                'latency_total_ms': qa.latency_total_ms,
                'tokens_total': qa.tokens_prompt + qa.tokens_completion,
            },
            'scores': [
                {
                    'dimension': s['dimension'],
                    'score': round(float(s['score'] or 0), 4),
                    'reason': s['reason'],
                    'eval_model': s['eval_model'],
                    'eval_latency_ms': s['eval_latency_ms'],
                    'eval_batch_id': s['eval_batch_id'],
                    'created_at': s['created_at'].isoformat(),
                }
                for s in scores
            ],
            'avg_score': round(avg_score, 4),
        })


# ============================================================================
# 低分归因分析 Views
# ============================================================================

class LowScoreAnalysisListView(APIView):
    """GET /api/v1/analytics/low-score-analysis/?days=7&category=&layer=&status=&root_type=&dept_id=&team_id=&limit=50

    低分归因列表:
    - 支持时间窗口、归因分类、影响层级、状态、组织归属筛选
    - select_related('qa_record') 避免 N+1(serializer 取 question/answer/root_type)
    - 默认按创建时间倒序,limit 上限 200 防止过大响应
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request):
        from apps.analytics.models import LowScoreAnalysis
        from apps.analytics.serializers import LowScoreAnalysisSerializer

        # 时间窗口
        try:
            days = int(request.query_params.get('days', 7))
        except (ValueError, TypeError):
            return Response({'detail': 'days 必须为整数'}, status=400)
        days = max(1, min(days, 90))
        since = timezone.now() - timedelta(days=days)

        try:
            limit = int(request.query_params.get('limit', 50))
        except (ValueError, TypeError):
            return Response({'detail': 'limit 必须为整数'}, status=400)
        limit = max(1, min(limit, 200))

        category = request.query_params.get('category', '').strip()
        layer = request.query_params.get('layer', '').strip()
        status = request.query_params.get('status', '').strip()
        root_type = request.query_params.get('root_type', '').strip()
        dept_id, team_id = _parse_org_scope(request)

        qs = LowScoreAnalysis.objects.select_related('qa_record').filter(created_at__gte=since)
        if category:
            qs = qs.filter(root_cause_category=category)
        if layer:
            qs = qs.filter(affected_layer=layer)
        if status:
            qs = qs.filter(status=status)
        if root_type:
            qs = qs.filter(qa_record__root_type=root_type)
        # 组织筛选:按 QaRecord.user 的归属 JOIN,LowScoreAnalysis 与 qa_record__user 关联
        qs = _apply_org_filter_on_qa(qs, dept_id, team_id, qa_prefix='qa_record__')

        qs = qs.order_by('-created_at')[:limit]
        rows = LowScoreAnalysisSerializer(qs, many=True).data
        return Response({'rows': rows, 'count': len(rows), 'days': days,
                         'dept_id': dept_id, 'team_id': team_id})


class LowScoreAnalysisDetailView(APIView):
    """GET /api/v1/analytics/low-score-analysis/?qa_record_id=123

    单条 QA 的归因详情(前端点击列表行展开用)。
    返回完整 question/answer + 归因结论 + 建议 + 低分维度 reason。
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request):
        from apps.analytics.models import LowScoreAnalysis
        from apps.analytics.serializers import LowScoreAnalysisSerializer

        qa_id = request.query_params.get('qa_record_id')
        if not qa_id:
            return Response({'detail': 'qa_record_id 必填'}, status=400)
        try:
            qa_id = int(qa_id)
        except (ValueError, TypeError):
            return Response({'detail': 'qa_record_id 必须为整数'}, status=400)

        try:
            analysis = LowScoreAnalysis.objects.select_related('qa_record').get(qa_record_id=qa_id)
        except LowScoreAnalysis.DoesNotExist:
            return Response({'detail': '该 QA 暂无归因分析,请先触发评估与归因'}, status=404)

        data = LowScoreAnalysisSerializer(analysis).data
        # 详情接口补充完整对话内容(列表已截断)
        data['full_question'] = analysis.qa_record.question if analysis.qa_record else ''
        data['full_answer'] = analysis.qa_record.answer if analysis.qa_record else ''
        return Response(data)


class RunLowScoreAnalysisView(APIView):
    """POST /api/v1/analytics/low-score-analysis/run/ - 手动触发单条 QA 归因

    场景:运营在看板发现某低分 QA,手动触发归因(跳过日预算,用户主动操作不计入生产配额)。
    异步执行:POST 立即返回,前端轮询 detail 接口查结果。
    若该 QA 尚未评估(无 MultiDimensionScore),返回 400 提示先评估。
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.write'

    def post(self, request):
        from apps.analytics.tasks import run_low_score_analysis
        from apps.analytics.models import MultiDimensionScore

        qa_id = request.data.get('qa_record_id')
        if not qa_id:
            return Response({'detail': 'qa_record_id 必填'}, status=400)
        try:
            qa_id = int(qa_id)
        except (ValueError, TypeError):
            return Response({'detail': 'qa_record_id 必须为整数'}, status=400)

        # 预检:无评估分数的 QA 无法归因,提前返回避免无意义排队
        has_scores = MultiDimensionScore.objects.filter(qa_record_id=qa_id).exists()
        if not has_scores:
            return Response({'detail': '该 QA 尚未评估,请先在「回答质量」执行 12 维评估'}, status=400)

        try:
            qa = QaRecord.objects.get(id=qa_id)
        except QaRecord.DoesNotExist:
            return Response({'detail': 'QA 记录不存在'}, status=404)

        threshold = request.data.get('threshold')
        if threshold is not None:
            try:
                threshold = float(threshold)
            except (TypeError, ValueError):
                return Response({'detail': 'threshold 必须为数字'}, status=400)

        # 派发异步归因任务(手动触发跳过日预算)
        run_low_score_analysis.delay(qa_id, threshold=threshold, skip_budget_check=True)
        logger.info(f'[LowScoreAnalysis] 手动触发归因 qa_id={qa_id} user={request.user.username}')

        return Response({
            'ok': True,
            'queued': True,
            'qa_id': qa_id,
            'message': '归因已派发,请通过 qa_record_id 轮询结果',
        })


class LowScoreAnalysisStatsView(APIView):
    """GET /api/v1/analytics/low-score-analysis/stats/?days=7&root_type=&dept_id=&team_id=

    归因分类统计(前端归因分布图用):
    - by_category: [{category, count, avg_score}]
    - by_layer: [{layer, count}]
    - by_method: {rule: n, llm: n, hybrid: n}
    - total: 总归因数
    一次 GROUP BY 查询拿全,避免逐类 COUNT
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request):
        from apps.analytics.models import LowScoreAnalysis

        try:
            days = int(request.query_params.get('days', 7))
        except (ValueError, TypeError):
            days = 7
        days = max(1, min(days, 90))
        root_type = request.query_params.get('root_type', '').strip()
        dept_id, team_id = _parse_org_scope(request)

        since = timezone.now() - timedelta(days=days)
        qs = LowScoreAnalysis.objects.filter(created_at__gte=since, status='completed')
        if root_type:
            qs = qs.filter(qa_record__root_type=root_type)
        qs = _apply_org_filter_on_qa(qs, dept_id, team_id, qa_prefix='qa_record__')

        # 按分类聚合:count + avg_score(一次 GROUP BY)
        cat_agg = qs.values('root_cause_category').annotate(
            count=models.Count('id'),
            avg_score=models.Avg('avg_score'),
        ).order_by('-count')
        by_category = [
            {
                'category': r['root_cause_category'],
                'count': r['count'],
                'avg_score': round(float(r['avg_score'] or 0), 4),
            }
            for r in cat_agg
        ]

        # 按层级聚合
        layer_agg = qs.values('affected_layer').annotate(
            count=models.Count('id'),
        ).order_by('-count')
        by_layer = [{'layer': r['affected_layer'], 'count': r['count']} for r in layer_agg]

        # 按方法聚合
        method_agg = qs.values('analysis_method').annotate(count=models.Count('id'))
        by_method = {r['analysis_method']: r['count'] for r in method_agg}

        total = sum(r['count'] for r in by_category)

        return Response({
            'days': days,
            'root_type': root_type or 'all',
            'dept_id': dept_id,
            'team_id': team_id,
            'total': total,
            'by_category': by_category,
            'by_layer': by_layer,
            'by_method': {
                'rule': by_method.get('rule', 0),
                'llm': by_method.get('llm', 0),
                'hybrid': by_method.get('hybrid', 0),
            },
        })


# ============================================================================
# 路由分析看板（LLM Wiki / GraphRAG / RAG 四层路由命中率与质量对比）
# ============================================================================

# 路由层级固定顺序：前端按此渲染堆叠图/柱状图，缺失的层补 0 而非跳过
ROUTE_ORDER = ['wiki', 'graphrag_local', 'graphrag_global', 'rag']
ROUTE_LABELS = {
    'wiki': 'Wiki 直答',
    'graphrag_local': 'GraphRAG 局部',
    'graphrag_global': 'GraphRAG 全局',
    'rag': 'RAG 兜底',
}


def _aggregate_query_transform_stats(qs):
    """从窗口内 QaRecord.route_trace 实时聚合查询改写/分解统计（改写命中率）

    改写链路不沉淀到 RouteAnalysis 表（主链 RAG 的 route_source 为空会被排除），
    故直接从 QaRecord.route_trace 统计：
    - rewrite_total: 走了改写链路的问答数（含 LLM 改写失败降级为原始 Query 的情况）
    - rewrite_changed: 其中 LLM 实际改写并改变了查询表述的次数
    - rewrite_hit_rate: rewrite_changed / rewrite_total，无改写链路时为 0
    - decompose_total: 触发查询分解的问答数（含分解失败降级的尝试）
    """
    rows = (qs.exclude(route_trace__isnull=True)
              .exclude(route_trace=[])
              .only('id', 'route_trace')
              .iterator())
    rewrite_total = rewrite_changed = decompose_total = 0
    for qa in rows:
        for entry in (qa.route_trace or []):
            layer = entry.get('layer')
            if layer == 'query_rewrite':
                rewrite_total += 1
                if entry.get('changed'):
                    rewrite_changed += 1
            elif layer == 'query_decompose':
                decompose_total += 1
    return {
        'rewrite_total': rewrite_total,
        'rewrite_changed': rewrite_changed,
        'rewrite_hit_rate': round(rewrite_changed / rewrite_total, 4) if rewrite_total else 0.0,
        'decompose_total': decompose_total,
    }


class RouteAnalysisDashboardView(APIView):
    """GET /api/v1/analytics/eval-dashboard/route-analysis/?days=7&dept_id=&team_id=

    路由分析看板：四层路由命中率 + 各维均分对比（由 aggregate_route_analysis_daily 任务供数）。

    - coverage_by_route: 每层命中数/占比/平均置信度/平均延迟/平均质量分
    - quality_by_route: 各层 12 维均分对比（按 4 大类分组，柱状/雷达图用）
    - daily_trend: 按天各层命中数（命中趋势堆叠图用）
    - query_transform_stats: 查询改写/分解统计（改写命中率，从 QaRecord.route_trace 实时聚合）

    时间窗口按 qa_created_at（提问时间）过滤；组织筛选按提问用户归属子查询
    （qa_record_id 为 BigInteger 非外键，无法直接 JOIN QaRecord，用子查询收敛）。
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request):
        from collections import defaultdict
        from apps.analytics.models import RouteAnalysis, MultiDimensionScore

        days = _parse_dashboard_days(request)
        dept_id, team_id = _parse_org_scope(request)
        since = timezone.now() - timedelta(days=days)

        qs = RouteAnalysis.objects.filter(qa_created_at__gte=since)
        if dept_id or team_id:
            # 组织筛选走 QaRecord.user 归属：子查询把命中 qa 收敛到组织内
            qa_ids_qs = _apply_org_filter_on_qa(QaRecord.objects.all(), dept_id, team_id)
            qs = qs.filter(qa_record_id__in=qa_ids_qs.values('id'))

        total = qs.count()
        if total == 0:
            return Response({
                'days': days,
                'dept_id': dept_id,
                'team_id': team_id,
                'total': 0,
                'route_order': ROUTE_ORDER,
                'route_labels': ROUTE_LABELS,
                'coverage_by_route': [],
                'quality_by_route': {},
                'daily_trend': [],
                'query_transform_stats': _aggregate_query_transform_stats(
                    _apply_org_filter_on_qa(QaRecord.objects.filter(created_at__gte=since),
                                            dept_id, team_id)),
            })

        # 1. 每层命中统计（一次 GROUP BY 拿全）
        route_agg = qs.values('route_source').annotate(
            count=models.Count('id'),
            avg_confidence=models.Avg('confidence'),
            avg_latency=models.Avg('latency_ms'),
            avg_quality=models.Avg('answer_quality'),
        ).order_by('-count')
        coverage_by_route = [
            {
                'route': r['route_source'],
                'count': r['count'],
                'share': round(r['count'] / total, 4),
                'avg_confidence': round(float(r['avg_confidence'] or 0), 4),
                'avg_latency_ms': round(float(r['avg_latency'] or 0), 1),
                'avg_answer_quality': round(float(r['avg_quality'] or 0), 4),
            }
            for r in route_agg
        ]

        # 2. 各层 12 维均分对比
        # 用子查询避免把窗口内 route_qa 全部拉进内存；再取 (qa_record_id -> route) 映射关联
        score_qs = MultiDimensionScore.objects.filter(
            qa_record_id__in=qs.values('qa_record_id'),
        )
        route_qa_map = dict(qs.values_list('qa_record_id', 'route_source'))

        dim_agg = score_qs.values('qa_record_id', 'dimension').annotate(
            avg_score=models.Avg('score'),
        )
        route_dim_avg = defaultdict(dict)  # route -> {dimension: avg}
        for r in dim_agg:
            route = route_qa_map.get(r['qa_record_id'])
            if route:
                route_dim_avg[route][r['dimension']] = round(float(r['avg_score'] or 0), 4)

        quality_by_route = {}
        for route in ROUTE_ORDER:
            dim_map = route_dim_avg.get(route, {})
            if not dim_map:
                quality_by_route[route] = {'overall': None, 'groups': {}, 'dimensions': {}}
                continue
            groups = {}
            for group_name, dims in _DIMENSION_GROUPS.items():
                group_avg = [dim_map[d] for d in dims if d in dim_map]
                if group_avg:
                    groups[group_name] = round(sum(group_avg) / len(group_avg), 4)
            # overall = 该层所有已评估维度均分的算术平均（与 4 大类组均一致）
            all_dims = list(dim_map.values())
            quality_by_route[route] = {
                'overall': round(sum(all_dims) / len(all_dims), 4) if all_dims else None,
                'groups': groups,
                'dimensions': dim_map,
            }

        # 3. 按天命中趋势（堆叠图数据）
        daily_qs = qs.annotate(
            _date=models.functions.TruncDate('qa_created_at'),
        ).values('_date', 'route_source').annotate(
            cnt=models.Count('id'),
        ).order_by('_date')
        trend_map = {}
        for r in daily_qs:
            date_str = r['_date'].isoformat()
            trend_map.setdefault(date_str, {})[r['route_source']] = r['cnt']
        daily_trend = []
        for date_str in sorted(trend_map.keys()):
            row = {'date': date_str}
            for route in ROUTE_ORDER:
                row[route] = trend_map[date_str].get(route, 0)
            daily_trend.append(row)

        return Response({
            'days': days,
            'dept_id': dept_id,
            'team_id': team_id,
            'total': total,
            'route_order': ROUTE_ORDER,
            'route_labels': ROUTE_LABELS,
            'coverage_by_route': coverage_by_route,
            'quality_by_route': quality_by_route,
            'daily_trend': daily_trend,
            # 查询改写/分解统计（改写命中率，与路由链路同窗口、同组织过滤）
            'query_transform_stats': _aggregate_query_transform_stats(
                _apply_org_filter_on_qa(QaRecord.objects.filter(created_at__gte=since),
                                        dept_id, team_id)),
        })


class RouteAnalysisAggregateView(APIView):
    """POST /api/v1/analytics/route-analysis/aggregate/ - 手动触发路由分析聚合

    body: {report_date: 'YYYY-MM-DD'} 可选；缺省聚合昨天。
    异步执行：POST 派发 Celery 任务立即返回，前端轮询看板数据刷新。
    用途：每日 beat 任务之外，运营改完路由配置后可立即重跑某天数据。
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.write'

    def post(self, request):
        from apps.analytics.tasks import aggregate_route_analysis_daily

        report_date = (request.data.get('report_date') or '').strip() or None
        if report_date:
            try:
                timezone.datetime.strptime(report_date, '%Y-%m-%d')
            except ValueError:
                return Response({'detail': 'report_date 格式须为 YYYY-MM-DD'}, status=400)

        aggregate_route_analysis_daily.delay(report_date)
        logger.info(
            f'[RouteAnalysis] 手动触发聚合 report_date={report_date or "yesterday"} '
            f'user={request.user.username}'
        )
        return Response({
            'ok': True,
            'queued': True,
            'report_date': report_date or 'yesterday',
            'message': '聚合已派发，稍后刷新看板即可看到结果',
        })


# ============================================================================
# Wiki 页面质量评估（忠实度 / 完整性 LLM-as-Judge）
# ============================================================================

class WikiQualityListView(APIView):
    """GET /api/v1/analytics/wiki-quality/?days=7&dimension=&status=&limit=&offset=&page_id=

    Wiki 页面质量评估结果列表：
    - summary: 统计（评估页数 / 失败页数 / 各维均分）
    - rows: 页面粒度明细（每页两个维度的 score + status + reason 截断）
    dimension 可选 faithfulness/completeness；status 可选 completed/failed；
    days 按评估更新时间窗口过滤；page_id 精确查某页（详情弹窗用，忽略窗口）。
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request):
        from apps.analytics.models import WikiPageQualityScore

        try:
            days = int(request.query_params.get('days', 7))
        except (ValueError, TypeError):
            days = 7
        days = max(1, min(days, 90))
        dimension = (request.query_params.get('dimension') or '').strip()
        status = (request.query_params.get('status') or '').strip()
        try:
            limit = int(request.query_params.get('limit', 50))
        except (ValueError, TypeError):
            limit = 50
        limit = max(1, min(limit, 200))
        try:
            offset = int(request.query_params.get('offset', 0))
        except (ValueError, TypeError):
            offset = 0
        offset = max(0, offset)

        since = timezone.now() - timedelta(days=days)
        qs = WikiPageQualityScore.objects.select_related('page').filter(updated_at__gte=since)
        if dimension:
            qs = qs.filter(dimension=dimension)
        if status:
            qs = qs.filter(status=status)
        # 精确查某页详情时(前端详情弹窗),忽略 days 窗口直接返回该页
        page_id = request.query_params.get('page_id', '').strip()
        if page_id:
            try:
                qs = qs.filter(page_id=int(page_id))
            except (ValueError, TypeError):
                return Response({'detail': 'page_id 必须为整数'}, status=400)

        # summary：各维均分 + 评估/失败页数（一次分组 + 两次 distinct 计数）
        dim_agg = qs.values('dimension').annotate(
            avg_score=models.Avg('score'),
            cnt=models.Count('id'),
        )
        summary = {'pages_evaluated': qs.values('page_id').distinct().count()}
        for r in dim_agg:
            summary[f'avg_{r["dimension"]}'] = round(float(r['avg_score'] or 0), 4)
            summary[f'count_{r["dimension"]}'] = r['cnt']
        summary['failed_pages'] = (
            qs.filter(status='failed').values('page_id').distinct().count()
        )

        # 明细：取窗口内 page_id 分页，再整批取每页两维分数（避免逐页 N+1）
        page_ids = list(
            qs.order_by('-updated_at').values_list('page_id', flat=True).distinct()[offset:offset + limit]
        )
        page_rows = {}
        if page_ids:
            score_rows = WikiPageQualityScore.objects.select_related('page').filter(
                page_id__in=page_ids,
            ).order_by('-updated_at')
            for s in score_rows:
                pm = page_rows.setdefault(s.page_id, {
                    'page_id': s.page_id,
                    'title': s.page.title,
                    'node_id': s.page.node_id,
                    'scores': {},
                })
                pm['scores'][s.dimension] = {
                    'score': round(float(s.score or 0), 4),
                    'status': s.status,
                    'reason': s.reason[:200],
                    'error_message': s.error_message[:200],
                    'updated_at': s.updated_at.isoformat(),
                }
        rows = list(page_rows.values())

        return Response({
            'days': days,
            'dimension': dimension or 'all',
            'status': status or 'all',
            'total': len(rows),
            'offset': offset,
            'limit': limit,
            'summary': summary,
            'rows': rows,
        })


class WikiQualityEvaluateView(APIView):
    """POST /api/v1/analytics/wiki-quality/evaluate/ - 手动触发 Wiki 页面质量批量评估

    body: {days: 7, limit: N} 可选；days 控制只重评估近期更新页面，limit 限制单次页面数。
    异步执行：POST 派发 Celery 任务立即返回，前端轮询 wiki-quality 列表查结果。
    目的：发布新 Wiki 或更新源文档后可手动触发，不必等每日 beat 任务。
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.write'

    def post(self, request):
        from apps.analytics.tasks import batch_evaluate_wiki_quality

        days = 7
        if request.data.get('days') is not None:
            try:
                days = int(request.data.get('days'))
            except (ValueError, TypeError):
                return Response({'detail': 'days 必须为整数'}, status=400)
            days = max(1, min(days, 90))

        limit = None
        if request.data.get('limit') is not None:
            try:
                limit = int(request.data.get('limit'))
            except (ValueError, TypeError):
                return Response({'detail': 'limit 必须为整数'}, status=400)
            limit = max(1, min(limit, 500))

        batch_evaluate_wiki_quality.delay(days=days, limit=limit)
        logger.info(
            f'[WikiEval] 手动触发批量评估 days={days} limit={limit} user={request.user.username}'
        )
        return Response({
            'ok': True,
            'queued': True,
            'days': days,
            'limit': limit,
            'message': '评估已派发，稍后刷新列表即可看到结果',
        })