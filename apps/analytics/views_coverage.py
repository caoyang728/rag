"""
analytics views - 覆盖率 & 反馈闭环
"""
import io

from loguru import logger

from django.http import HttpResponse
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.users.permissions import CanViewAnalytics

# ============================================================================
# 覆盖率 & 反馈闭环 Views
# ============================================================================

class CoverageReportView(APIView):
    """GET /api/v1/analytics/coverage/ - 覆盖率报告"""
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request):
        from apps.analytics.services.coverage_service import (
            analyze_hot_query_coverage, detect_knowledge_gaps,
            detect_duplicate_chunks, analyze_domain_coverage,
        )
        try:
            days = int(request.query_params.get('days', 7))
        except (ValueError, TypeError):
            return Response({'detail': 'days 必须为整数'}, status=400)
        # 限制 days 范围：1-30 天，防止过大范围扫描全表
        days = max(1, min(days, 30))

        coverage = analyze_hot_query_coverage(days)
        gaps = detect_knowledge_gaps(days)
        duplicates = detect_duplicate_chunks()
        domain = analyze_domain_coverage(days)

        return Response({
            'coverage': coverage,
            'gaps': gaps[:20],
            'gap_count': len(gaps),
            'duplicates': duplicates,
            'domain': domain,
        })


class FeedbackLoopView(APIView):
    """POST /api/v1/analytics/feedback-loop/ - 执行反馈闭环分析"""
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.write'

    def post(self, request):
        from apps.analytics.services.coverage_service import auto_link_feedback_to_chunks
        try:
            days = int(request.data.get('days', 7))
        except (ValueError, TypeError):
            return Response({'detail': 'days 必须为整数'}, status=400)
        days = max(1, min(days, 30))
        result = auto_link_feedback_to_chunks(days=days)
        return Response(result)


class GenerateCoverageReportView(APIView):
    """POST /api/v1/analytics/coverage/generate/ - 生成覆盖率报告"""
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.write'

    def post(self, request):
        from apps.analytics.services.coverage_service import generate_coverage_report
        try:
            days = int(request.data.get('days', 7))
        except (ValueError, TypeError):
            return Response({'detail': 'days 必须为整数'}, status=400)
        days = max(1, min(days, 30))
        report = generate_coverage_report(days=days)
        return Response({
            'ok': True,
            'report_id': report.id,
            'report_date': str(report.report_date),
            'coverage_rate': report.hot_query_coverage_rate,
            'gap_count': report.gap_count,
        })


class CoverageReportListView(APIView):
    """GET /api/v1/analytics/coverage/reports/ - 历史覆盖率报告列表

    - 返回最近 50 条报告记录，按日期倒序
    - 供前端「历史报告」面板展示，支持下载和删除
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request):
        from apps.analytics.models import CoverageReport
        # 最近 50 条，避免全表返回
        rows = list(
            CoverageReport.objects
            .order_by('-report_date', '-created_at')
            .values(
                'id', 'report_date', 'total_hot_queries', 'covered_queries',
                'uncovered_queries', 'hot_query_coverage_rate', 'gap_count',
                'duplicate_chunk_rate', 'duplicate_chunk_count',
                'feedback_loop_count', 'feedback_resolved_count',
                'created_at',
            )[:50]
        )
        return Response({'rows': rows, 'count': len(rows)})


class CoverageReportDetailView(APIView):
    """DELETE /api/v1/analytics/coverage/reports/<id>/ - 删除覆盖率报告

    - 仅允许删除历史报告，不影响当前覆盖率展示
    - 删除操作记录审计日志
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.write'

    def delete(self, request, report_id):
        from apps.analytics.models import CoverageReport
        try:
            report = CoverageReport.objects.get(id=report_id)
        except CoverageReport.DoesNotExist:
            return Response({'detail': '报告不存在'}, status=404)
        report.delete()
        logger.info(
            f'coverage_report_deleted report_id={report_id} date={report.report_date} user={request.user.username}'
        )
        return Response({'ok': True})


class CoverageReportExportView(APIView):
    """GET /api/v1/analytics/coverage/reports/<id>/export/ - 导出覆盖率报告为 Excel

    - 将单条报告的完整数据导出为 .xlsx 文件
    - 包含：覆盖率概览、知识空白详情、部门/团队覆盖明细
    - 使用 openpyxl 生成多 Sheet Excel
    """
    permission_classes = [IsAuthenticated, CanViewAnalytics]
    required_perm = 'analytics.system.read'

    def get(self, request, report_id):
        from apps.analytics.models import CoverageReport
        try:
            report = CoverageReport.objects.get(id=report_id)
        except CoverageReport.DoesNotExist:
            return Response({'detail': '报告不存在'}, status=404)

        # 使用 openpyxl 生成 Excel，比 csv 支持多 Sheet + 格式化
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()

        # ===== Sheet 1: 覆盖率概览 =====
        ws1 = wb.active
        ws1.title = '覆盖率概览'
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')

        overview = [
            ('报告ID', report.id),
            ('报告日期', str(report.report_date)),
            ('生成时间', report.created_at.strftime('%Y-%m-%d %H:%M')),
            ('', ''),
            ('热门查询总数', report.total_hot_queries),
            ('已覆盖查询数', report.covered_queries),
            ('未覆盖查询数', report.uncovered_queries),
            ('热门问题覆盖率', f'{report.hot_query_coverage_rate * 100:.1f}%'),
            ('知识空白数', report.gap_count),
            ('重复切片率', f'{report.duplicate_chunk_rate * 100:.1f}%'),
            ('重复切片数', report.duplicate_chunk_count),
            ('反馈关联数', report.feedback_loop_count),
            ('反馈已解决数', report.feedback_resolved_count),
        ]
        for row_idx, (label, value) in enumerate(overview, 1):
            ws1.cell(row=row_idx, column=1, value=label).font = header_font
            ws1.cell(row=row_idx, column=1).fill = header_fill
            ws1.cell(row=row_idx, column=2, value=value)
        ws1.column_dimensions['A'].width = 20
        ws1.column_dimensions['B'].width = 30

        # ===== Sheet 2: 知识空白详情 =====
        ws2 = wb.create_sheet('知识空白')
        gaps = report.gap_queries or []
        ws2.append(['查询内容', '出现次数', '改进建议'])
        for col in range(1, 4):
            cell = ws2.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for g in gaps:
            ws2.append([
                g.get('query', ''),
                g.get('count', 0),
                g.get('suggestion', ''),
            ])
        ws2.column_dimensions['A'].width = 40
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 50

        # ===== Sheet 3: 部门/团队覆盖明细 =====
        ws3 = wb.create_sheet('部门覆盖')
        domain = report.domain_coverage or {}
        domain_list = domain.get('domain_coverage', []) if isinstance(domain, dict) else domain
        ws3.append(['部门/团队', '文档数', '切片数', '占比', '命中率', '下属团队数'])
        for col in range(1, 7):
            cell = ws3.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for d in domain_list:
            teams = d.get('teams', []) if isinstance(d, dict) else []
            # 兼容 dict 和 list 两种数据格式
            doc_count = d.get('doc_count', 0) if isinstance(d, dict) else 0
            chunk_count = d.get('chunk_count', 0) if isinstance(d, dict) else 0
            share = d.get('占比', 0) if isinstance(d, dict) else 0
            hit_rate = d.get('query_hit_rate', 0) if isinstance(d, dict) else 0
            name = d.get('name', '') if isinstance(d, dict) else str(d)
            ws3.append([
                name,
                doc_count,
                chunk_count,
                f'{share * 100:.1f}%' if isinstance(share, (int, float)) else str(share),
                f'{hit_rate * 100:.1f}%' if isinstance(hit_rate, (int, float)) else str(hit_rate),
                len(teams),
            ])
            # 如果有子团队，在下一行缩进展示
            if isinstance(teams, list):
                for team_item in teams:
                    if isinstance(team_item, (list, tuple)) and len(team_item) == 2:
                        team_name, team_data = team_item
                        t_docs = team_data.get('doc_count', 0) if isinstance(team_data, dict) else 0
                        t_chunks = team_data.get('chunk_count', 0) if isinstance(team_data, dict) else 0
                        ws3.append([f'  └ {team_name}', t_docs, t_chunks, '', '', ''])
        ws3.column_dimensions['A'].width = 25
        for col in 'BCDEF':
            ws3.column_dimensions[col].width = 12

        # 输出为 HTTP 响应，浏览器直接下载
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'coverage_report_{report.report_date}.xlsx'
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        logger.info(
            f'coverage_report_exported report_id={report_id} user={request.user.username}'
        )
        return response
