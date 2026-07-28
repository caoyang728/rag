"""
电子表格解析器（CSV / XLSX / XLS / WPS ET）
- CSV: 使用标准库 csv 模块，自动检测编码
- XLSX: 使用 openpyxl 读取，保留 sheet 结构
- XLS/ET: 旧版二进制格式，尝试 openpyxl 解析，失败则降级为文本提取
每个 sheet 输出为一个 table block，保留 sheet 名称作为 section_path
"""
import csv
import os
from loguru import logger
from typing import List, Dict, Any

from .base import BaseParser


class SpreadsheetParser(BaseParser):
    name = 'spreadsheet'

    def parse(self, file_path: str, **options) -> List[Dict[str, Any]]:
        ext = os.path.splitext(file_path)[1].lower()

        if ext == '.csv':
            return self._parse_csv(file_path)
        elif ext in ('.xlsx', '.et'):
            return self._parse_xlsx(file_path)
        elif ext == '.xls':
            # .xls 是旧版二进制格式，openpyxl 不支持，尝试降级处理
            logger.warning(f'[SpreadsheetParser] .xls 旧格式支持有限，尝试降级解析: {file_path}')
            return self._parse_xls_fallback(file_path)
        else:
            # 未知扩展名，尝试 xlsx 方式
            return self._parse_xlsx(file_path)

    def _parse_csv(self, file_path: str) -> List[Dict[str, Any]]:
        """解析 CSV 文件，自动检测编码"""
        blocks: List[Dict[str, Any]] = []

        # 尝试多种编码
        content = None
        for encoding in ('utf-8', 'gbk', 'gb2312', 'latin-1'):
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    content = f.read()
                break
            except (UnicodeDecodeError, Exception):
                continue

        if content is None:
            logger.error(f'[SpreadsheetParser] CSV 编码检测失败: {file_path}')
            return []

        try:
            # 自动检测分隔符
            sample = content[:1024]
            dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
            reader = csv.reader(content.splitlines(), dialect)
        except csv.Error:
            # 检测失败，默认用逗号
            reader = csv.reader(content.splitlines())

        rows = list(reader)
        if not rows:
            return []

        # 转为 markdown 表格格式
        lines = []
        for i, row in enumerate(rows):
            # 转义单元格中的管道符
            cells = [str(cell).replace('|', '\\|').replace('\n', ' ').strip() for cell in row]
            lines.append(' | '.join(cells))
            if i == 0:
                # 表头后加分隔行
                lines.append(' | '.join(['---'] * len(row)))

        table_content = '\n'.join(lines)
        blocks.append({
            'type': 'table',
            'content': table_content[:50000],
            'section_path': 'CSV 数据',
            'page_number': None,
            'extra': {
                'rows': len(rows),
                'cols': len(rows[0]) if rows else 0,
                'format': 'csv',
            },
        })
        return blocks

    def _parse_xlsx(self, file_path: str) -> List[Dict[str, Any]]:
        """使用 openpyxl 解析 XLSX/ET 文件"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.error('[SpreadsheetParser] openpyxl 未安装，无法解析 XLSX')
            return []

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception:
            logger.exception(f'[SpreadsheetParser] 打开 XLSX 失败: {file_path}')
            return []

        blocks: List[Dict[str, Any]] = []
        for ws in wb.worksheets:
            rows_data = []
            for row in ws.iter_rows(values_only=True):
                # 将 None 转为空字符串，其他转为字符串
                cells = [str(cell) if cell is not None else '' for cell in row]
                # 跳过全空行
                if any(c.strip() for c in cells):
                    rows_data.append(cells)

            if not rows_data:
                continue

            # 转为 markdown 表格
            lines = []
            for i, row_cells in enumerate(rows_data):
                escaped = [c.replace('|', '\\|').replace('\n', ' ').strip() for c in row_cells]
                lines.append(' | '.join(escaped))
                if i == 0:
                    lines.append(' | '.join(['---'] * len(row_cells)))

            table_content = '\n'.join(lines)
            blocks.append({
                'type': 'table',
                'content': table_content[:50000],
                'section_path': f'Sheet: {ws.title}',
                'page_number': None,
                'extra': {
                    'rows': len(rows_data),
                    'cols': len(rows_data[0]) if rows_data else 0,
                    'sheet_name': ws.title,
                    'format': 'xlsx',
                },
            })

        wb.close()
        return blocks

    def _parse_xls_fallback(self, file_path: str) -> List[Dict[str, Any]]:
        """.xls 旧版二进制格式降级处理：尝试用 openpyxl（会失败），然后提示不支持"""
        # openpyxl 不支持 .xls，记录日志并返回提示信息
        logger.warning(f'[SpreadsheetParser] .xls 旧版格式无法解析，建议转换为 .xlsx: {file_path}')
        return [{
            'type': 'text',
            'content': '[此文件为旧版 .xls 格式，系统暂不支持自动解析。请将文件另存为 .xlsx 格式后重新上传。]',
            'section_path': '解析提示',
            'page_number': None,
            'extra': {'format': 'xls', 'parse_error': 'unsupported_legacy_format'},
        }]
